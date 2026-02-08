#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
QA DB Query Script for FMEA Analysis
QA_품질이력.db에서 FMEA 작성에 필요한 데이터를 추출

Usage:
    python query_qa_db.py <db_path> [--component <부품명>] [--mode <고장형태>]
    python query_qa_db.py <db_path> --stats
    python query_qa_db.py <db_path> --export <output.xlsx>

Examples:
    python query_qa_db.py QA_품질이력.db --component 권선
    python query_qa_db.py QA_품질이력.db --stats
    python query_qa_db.py QA_품질이력.db --export fmea_mapping.xlsx

Requirements:
    pip install openpyxl pandas

Author: Claude (FMEA Analysis Skill)
Version: 1.0
Date: 2025-01-27
"""

import sys
import io
import sqlite3
import argparse
from pathlib import Path
from collections import defaultdict
from datetime import datetime

# Windows cp949 인코딩 문제 해결
if sys.stdout:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# FMEA 컬럼 매핑 (qa-data-mapping.md 기반)
FMEA_COLUMN_MAPPING = {
    # A열: 부품명
    'A_component': ['품명'],

    # B열: 기능 (추론 필요)
    'B_function': ['품명', '항목구분'],

    # C열: 고장영향
    'C_effect': ['발생현상유형', '발생현상유형소분류', '현상_소분류'],

    # D열: S값 (심각도)
    'D_severity': ['중요_경미', '치명도', '분류', '피해보상비'],

    # E열: 고장형태
    'E_failure_mode': ['발생현상유형', '발생현상유형소분류', '현상_소분류'],

    # F열: 고장원인
    'F_cause': ['발생원인', '발생원인유형', '원인부서'],

    # G열: O값 (발생도)
    'G_occurrence': ['발생년도', '품명', 'COUNT'],

    # H열: 현 예방관리
    'H_prevention': ['조치내역', '조치세부내역'],

    # I열: D값 (검출도)
    'I_detection': ['검사구분', '항목구분', '검사소분류'],

    # J열: 현 검출관리
    'J_detection_control': ['검사구분', '항목구분'],

    # K열: AP (조치우선순위)
    'K_ap': ['중요_경미', '분류'],
}

# S값 자동 계산 기준
S_VALUE_MAPPING = {
    '중요_경미': {
        '중요': 8,
        '경미': 4,
    },
    '치명도': {
        'A': 10,
        'B': 8,
        'C': 6,
        'D': 4,
    },
    '분류_weight': {
        'CLAIM': 1,  # +1 가중치
        'NCR': 0,
    },
    '피해보상비_threshold': [
        (100_000_000, 10),  # 1억 이상: S=10
        (10_000_000, 1),    # 1천만 이상: +1
        (0, 0),
    ],
}

# D값 자동 계산 기준
D_VALUE_MAPPING = {
    '검사구분': {
        '수입검사': 6,
        '공정검사': 5,
        '최종검사': 4,
        '출하검사': 3,
        '설치검사': 4,
    },
    '항목구분': {
        '외관검사': 5,
        '치수검사': 4,
        '기능검사': 3,
        '시험검사': 2,
        '성능시험': 2,
    },
}

# F열 라이프사이클 태그 매핑
LIFECYCLE_TAG_MAPPING = {
    '설계': ['설계', '도면', '규격', '사양', 'CAD'],
    '재료': ['자재', '재료', '부품', '원자재', '외주품'],
    '제작': ['가공', '조립', '제조', '용접', '코일', '제작'],
    '시험': ['시험', '검사', '측정', '출하', '완성'],
}


def connect_db(db_path):
    """SQLite DB 연결"""
    if not Path(db_path).exists():
        print(f"[ERROR] DB file not found: {db_path}")
        sys.exit(1)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def get_all_columns(conn):
    """모든 컬럼 목록 조회"""
    cursor = conn.execute("PRAGMA table_info(qa_records)")
    columns = [row[1] for row in cursor.fetchall()]
    return columns


def query_by_component(conn, component_name):
    """부품명으로 QA 이력 조회"""
    query = """
    SELECT
        품명,
        발생현상유형,
        발생현상유형소분류,
        현상_소분류,
        발생원인,
        발생원인유형,
        원인부서,
        중요_경미,
        치명도,
        분류,
        피해보상비,
        검사구분,
        항목구분,
        조치내역,
        발생년도,
        COUNT(*) as 발생횟수
    FROM qa_records
    WHERE 품명 LIKE ?
    GROUP BY
        품명, 발생현상유형, 발생현상유형소분류,
        발생원인, 발생원인유형, 중요_경미, 치명도
    ORDER BY 발생횟수 DESC
    """
    cursor = conn.execute(query, (f'%{component_name}%',))
    return cursor.fetchall()


def query_by_failure_mode(conn, failure_mode):
    """고장형태로 QA 이력 조회"""
    query = """
    SELECT
        품명,
        발생현상유형,
        발생현상유형소분류,
        현상_소분류,
        발생원인,
        발생원인유형,
        원인부서,
        중요_경미,
        치명도,
        분류,
        피해보상비,
        검사구분,
        항목구분,
        조치내역,
        발생년도,
        COUNT(*) as 발생횟수
    FROM qa_records
    WHERE 발생현상유형 LIKE ? OR 발생현상유형소분류 LIKE ?
    GROUP BY
        품명, 발생현상유형, 발생현상유형소분류,
        발생원인, 발생원인유형
    ORDER BY 발생횟수 DESC
    """
    cursor = conn.execute(query, (f'%{failure_mode}%', f'%{failure_mode}%'))
    return cursor.fetchall()


def calc_s_value(row):
    """S값 자동 계산"""
    base = 5  # 기본값

    # 중요/경미 기준
    if row['중요_경미']:
        base = S_VALUE_MAPPING['중요_경미'].get(row['중요_경미'], 5)

    # 치명도 기준 (있으면 우선)
    if row['치명도']:
        치명도_s = S_VALUE_MAPPING['치명도'].get(row['치명도'])
        if 치명도_s:
            base = 치명도_s

    # CLAIM 가중치
    if row['분류'] == 'CLAIM':
        base += 1

    # 피해보상비 가중치
    cost = row['피해보상비'] or 0
    if isinstance(cost, str):
        try:
            cost = int(cost.replace(',', ''))
        except:
            cost = 0

    if cost >= 100_000_000:
        return 10
    elif cost >= 10_000_000:
        base += 1

    return min(base, 10)


def calc_d_value(row):
    """D값 자동 계산"""
    base = 5  # 기본값

    # 검사구분 기준
    if row['검사구분']:
        base = D_VALUE_MAPPING['검사구분'].get(row['검사구분'], 5)

    # 항목구분으로 조정
    if row['항목구분']:
        adjustment = D_VALUE_MAPPING['항목구분'].get(row['항목구분'])
        if adjustment:
            base = min(base, adjustment)

    return base


def get_lifecycle_tag(row):
    """F열 라이프사이클 태그 추론"""
    tags = set()

    # 발생원인유형 기반
    cause_type = row['발생원인유형'] or ''
    for tag, keywords in LIFECYCLE_TAG_MAPPING.items():
        for kw in keywords:
            if kw in cause_type:
                tags.add(tag)
                break

    # 원인부서 기반
    dept = row['원인부서'] or ''
    if '설계' in dept:
        tags.add('설계')
    elif '자재' in dept or '구매' in dept:
        tags.add('재료')
    elif '생산' in dept or '제조' in dept or '가공' in dept:
        tags.add('제작')
    elif '품질' in dept or '검사' in dept or '시험' in dept:
        tags.add('시험')

    return list(tags) if tags else ['제작']  # 기본값


def get_statistics(conn):
    """QA DB 통계 정보"""
    stats = {}

    # 총 레코드 수
    cursor = conn.execute("SELECT COUNT(*) FROM qa_records")
    stats['total_records'] = cursor.fetchone()[0]

    # 분류별 통계
    cursor = conn.execute("""
        SELECT 분류, COUNT(*) as cnt
        FROM qa_records
        GROUP BY 분류
        ORDER BY cnt DESC
    """)
    stats['by_category'] = dict(cursor.fetchall())

    # 발생현상유형 상위 10
    cursor = conn.execute("""
        SELECT 발생현상유형, COUNT(*) as cnt
        FROM qa_records
        WHERE 발생현상유형 IS NOT NULL AND 발생현상유형 != ''
        GROUP BY 발생현상유형
        ORDER BY cnt DESC
        LIMIT 10
    """)
    stats['top_failure_types'] = list(cursor.fetchall())

    # 발생원인유형 상위 10
    cursor = conn.execute("""
        SELECT 발생원인유형, COUNT(*) as cnt
        FROM qa_records
        WHERE 발생원인유형 IS NOT NULL AND 발생원인유형 != ''
        GROUP BY 발생원인유형
        ORDER BY cnt DESC
        LIMIT 10
    """)
    stats['top_cause_types'] = list(cursor.fetchall())

    # 연도별 추이
    cursor = conn.execute("""
        SELECT 발생년도, COUNT(*) as cnt
        FROM qa_records
        WHERE 발생년도 IS NOT NULL
        GROUP BY 발생년도
        ORDER BY 발생년도 DESC
    """)
    stats['by_year'] = list(cursor.fetchall())

    # 품명별 통계 상위 10
    cursor = conn.execute("""
        SELECT 품명, COUNT(*) as cnt
        FROM qa_records
        WHERE 품명 IS NOT NULL AND 품명 != ''
        GROUP BY 품명
        ORDER BY cnt DESC
        LIMIT 10
    """)
    stats['top_components'] = list(cursor.fetchall())

    return stats


def export_fmea_mapping(conn, output_path):
    """FMEA 매핑 데이터 Excel 내보내기"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[ERROR] openpyxl not installed. Run: pip install openpyxl")
        return

    wb = Workbook()

    # Sheet 1: FMEA 매핑 데이터
    ws = wb.active
    ws.title = "FMEA_Mapping"

    headers = [
        'A_부품명', 'C_고장영향_대분류', 'C_고장영향_소분류', 'C_고장영향_상세',
        'D_S값', 'D_중요경미', 'D_치명도', 'D_분류', 'D_피해보상비',
        'E_고장형태', 'F_고장원인', 'F_원인유형', 'F_라이프사이클',
        'I_D값', 'I_검사구분', 'I_항목구분',
        '발생년도', '발생횟수'
    ]

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 15

    # 데이터 조회
    query = """
    SELECT
        품명,
        발생현상유형,
        발생현상유형소분류,
        현상_소분류,
        중요_경미,
        치명도,
        분류,
        피해보상비,
        발생원인,
        발생원인유형,
        원인부서,
        검사구분,
        항목구분,
        발생년도,
        COUNT(*) as 발생횟수
    FROM qa_records
    GROUP BY
        품명, 발생현상유형, 발생현상유형소분류, 현상_소분류,
        발생원인, 발생원인유형, 원인부서,
        중요_경미, 치명도, 분류, 검사구분, 항목구분, 발생년도
    ORDER BY 발생횟수 DESC
    """
    cursor = conn.execute(query)

    row_num = 2
    for row in cursor.fetchall():
        row_dict = dict(row)

        # S값 계산
        s_value = calc_s_value(row_dict)

        # D값 계산
        d_value = calc_d_value(row_dict)

        # 라이프사이클 태그
        lifecycle = ', '.join(get_lifecycle_tag(row_dict))

        data = [
            row_dict['품명'],
            row_dict['발생현상유형'],
            row_dict['발생현상유형소분류'],
            row_dict['현상_소분류'],
            s_value,
            row_dict['중요_경미'],
            row_dict['치명도'],
            row_dict['분류'],
            row_dict['피해보상비'],
            row_dict['발생현상유형소분류'] or row_dict['발생현상유형'],  # 고장형태
            row_dict['발생원인'],
            row_dict['발생원인유형'],
            lifecycle,
            d_value,
            row_dict['검사구분'],
            row_dict['항목구분'],
            row_dict['발생년도'],
            row_dict['발생횟수'],
        ]

        for col, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col, value=value)

        row_num += 1

    # Sheet 2: 통계 요약
    ws_stats = wb.create_sheet(title="Statistics")
    stats = get_statistics(conn)

    ws_stats.cell(row=1, column=1, value="QA DB Statistics")
    ws_stats.cell(row=1, column=1).font = Font(bold=True, size=14)

    ws_stats.cell(row=3, column=1, value="Total Records")
    ws_stats.cell(row=3, column=2, value=stats['total_records'])

    ws_stats.cell(row=5, column=1, value="By Category (분류)")
    row = 6
    for cat, cnt in stats['by_category'].items():
        ws_stats.cell(row=row, column=1, value=cat)
        ws_stats.cell(row=row, column=2, value=cnt)
        row += 1

    ws_stats.cell(row=row+1, column=1, value="Top 10 Failure Types (발생현상유형)")
    row += 2
    for ft, cnt in stats['top_failure_types']:
        ws_stats.cell(row=row, column=1, value=ft)
        ws_stats.cell(row=row, column=2, value=cnt)
        row += 1

    ws_stats.cell(row=row+1, column=1, value="Top 10 Cause Types (발생원인유형)")
    row += 2
    for ct, cnt in stats['top_cause_types']:
        ws_stats.cell(row=row, column=1, value=ct)
        ws_stats.cell(row=row, column=2, value=cnt)
        row += 1

    # 저장
    wb.save(output_path)
    print(f"[OK] FMEA mapping exported: {output_path}")
    print(f"   - Total rows: {row_num - 2}")
    print(f"   - Sheets: FMEA_Mapping, Statistics")


def print_results(results, title):
    """결과 출력"""
    print(f"\n{'='*70}")
    print(f" {title}")
    print(f"{'='*70}")
    print(f" Total: {len(results)} records\n")

    if not results:
        print(" No data found.")
        return

    for i, row in enumerate(results[:20], 1):  # 상위 20개만
        row_dict = dict(row)
        s_val = calc_s_value(row_dict)
        d_val = calc_d_value(row_dict)
        lifecycle = get_lifecycle_tag(row_dict)

        print(f" [{i:2d}] 품명: {row_dict['품명']}")
        print(f"      고장영향: {row_dict['발생현상유형']} > {row_dict['발생현상유형소분류']}")
        print(f"      고장원인: {row_dict['발생원인']} ({row_dict['발생원인유형']})")
        print(f"      S값: {s_val} (중요경미:{row_dict['중요_경미']}, 치명도:{row_dict['치명도']}, 분류:{row_dict['분류']})")
        print(f"      D값: {d_val} (검사:{row_dict['검사구분']}, 항목:{row_dict['항목구분']})")
        print(f"      라이프사이클: {', '.join(lifecycle)}")
        print(f"      발생횟수: {row_dict['발생횟수']}회")
        print()

    if len(results) > 20:
        print(f" ... and {len(results) - 20} more records")


def print_statistics(stats):
    """통계 출력"""
    print(f"\n{'='*70}")
    print(f" QA DB Statistics")
    print(f"{'='*70}")

    print(f"\n [Total Records] {stats['total_records']:,}")

    print(f"\n [By Category (분류)]")
    for cat, cnt in stats['by_category'].items():
        print(f"   - {cat}: {cnt:,}")

    print(f"\n [Top 10 Failure Types (발생현상유형)]")
    for ft, cnt in stats['top_failure_types']:
        print(f"   - {ft}: {cnt:,}")

    print(f"\n [Top 10 Cause Types (발생원인유형)]")
    for ct, cnt in stats['top_cause_types']:
        print(f"   - {ct}: {cnt:,}")

    print(f"\n [By Year (발생년도)]")
    for year, cnt in stats['by_year'][:5]:
        print(f"   - {year}: {cnt:,}")

    print(f"\n [Top 10 Components (품명)]")
    for comp, cnt in stats['top_components']:
        print(f"   - {comp}: {cnt:,}")


def main():
    parser = argparse.ArgumentParser(
        description='Query QA DB for FMEA analysis',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python query_qa_db.py QA_품질이력.db --component 권선
  python query_qa_db.py QA_품질이력.db --mode 절연파괴
  python query_qa_db.py QA_품질이력.db --stats
  python query_qa_db.py QA_품질이력.db --export fmea_mapping.xlsx
        """
    )
    parser.add_argument('db_path', help='Path to QA_품질이력.db')
    parser.add_argument('--component', '-c', help='Filter by component name (품명)')
    parser.add_argument('--mode', '-m', help='Filter by failure mode (고장형태)')
    parser.add_argument('--stats', '-s', action='store_true', help='Show statistics')
    parser.add_argument('--export', '-e', help='Export to Excel file')
    parser.add_argument('--columns', action='store_true', help='List all columns')

    args = parser.parse_args()

    conn = connect_db(args.db_path)

    try:
        if args.columns:
            columns = get_all_columns(conn)
            print(f"\n[QA DB Columns] Total: {len(columns)}")
            for i, col in enumerate(columns, 1):
                print(f"  {i:2d}. {col}")

        elif args.stats:
            stats = get_statistics(conn)
            print_statistics(stats)

        elif args.export:
            export_fmea_mapping(conn, args.export)

        elif args.component:
            results = query_by_component(conn, args.component)
            print_results(results, f"QA Records for Component: {args.component}")

        elif args.mode:
            results = query_by_failure_mode(conn, args.mode)
            print_results(results, f"QA Records for Failure Mode: {args.mode}")

        else:
            # 기본: 통계 출력
            stats = get_statistics(conn)
            print_statistics(stats)
            print(f"\n[TIP] Use --help for more options")

    finally:
        conn.close()


if __name__ == "__main__":
    main()
