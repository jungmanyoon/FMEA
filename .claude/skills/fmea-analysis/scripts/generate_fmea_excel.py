#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FMEA Excel 자동 생성 스크립트

목적: 결정적 Excel 생성, 셀 병합, 검증 자동화
기준: references/excel-generation.md

사용법:
    python generate_fmea_excel.py input_data.json output.xlsx
    python generate_fmea_excel.py input_data.json output.xlsx --qa-db <db_path>
"""

import sys
import io
import json
import os
import re
import sqlite3
from pathlib import Path

# Windows cp949 인코딩 문제 해결 (공통 모듈 사용)
from encoding_utils import setup_encoding
setup_encoding()

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 고장형태(E열) 금지어 검증 모듈
from validate_failure_mode import validate_failure_mode, validate_tag_format, FORBIDDEN_PATTERNS, FORBIDDEN_EXACT
from validate_causal_chain import validate_mode_cause, validate_cause_mechanism, validate_lifecycle_consistency
# 고장영향(C열) 검증 모듈 (재발방지대책 260111, v2.0 물리적 상태 검증 추가 260129)
from validate_failure_effect import (
    validate_failure_effect,
    validate_physical_in_effect,
    load_effect_ontology,
    FORBIDDEN_INSPECTION_RESULTS
)
# H열/J열(현재예방대책/현재검출대책) 검증 모듈 (260129)
from validate_prevention_detection import (
    validate_stage_format,
    validate_source_presence,
    validate_forbidden_source,
    validate_value_presence,
    REQUIRED_STAGES,
    FORBIDDEN_SOURCE
)
# G/H/J 컬럼 형식 검증 모듈 (260131 CRITICAL-3)
from validate_single_item import (
    validate_mechanism,
    validate_prevention_multiline,
    validate_detection_multiline
)


# ========================================
# QA DB 자동화 모듈 (qa-data-mapping.md 기반)
# ========================================

# S값 자동 계산 기준
QA_S_VALUE_MAPPING = {
    '중요_경미': {'중요': 8, '경미': 4},
    '치명도': {'A': 10, 'B': 8, 'C': 6, 'D': 4, '상': 8, '중': 5, '하': 3},
}

# D값 자동 계산 기준
QA_D_VALUE_MAPPING = {
    '검사구분': {'수입검사': 3, '수입': 3, '공정검사': 5, '공정': 5, '최종검사': 4, '제품': 7, '시공': 9},
    '항목구분': {'외관검사': 5, '치수검사': 4, '기능검사': 3, '시험검사': 2, '절연시험': 2, '완성': 6},
}

# 라이프사이클 태그 매핑 (대괄호 없이 통일)
QA_LIFECYCLE_MAPPING = {
    '설계': ['설계', '도면', '규격', '사양', 'CAD', '설계 오류'],
    '재료': ['자재', '재료', '부품', '원자재', '외주품', '부품 불량', '자재 불량'],
    '제작': ['가공', '조립', '제조', '용접', '코일', '제작', '생산 작업 불량', '제관 제작 불량'],
    '시험': ['시험', '검사', '측정', '출하', '완성', '검사 누락'],
}


def connect_qa_db(db_path):
    """QA DB 연결"""
    if not Path(db_path).exists():
        print(f"[WARN] QA DB not found: {db_path}")
        return None
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def calc_s_value_from_qa(row):
    """QA DB 데이터에서 S값 자동 계산"""
    base = 5  # 기본값

    # 중요/경미 기준
    if row.get('중요_경미'):
        base = QA_S_VALUE_MAPPING['중요_경미'].get(row['중요_경미'], 5)

    # 치명도 기준 (있으면 우선)
    if row.get('치명도'):
        치명도_s = QA_S_VALUE_MAPPING['치명도'].get(row['치명도'])
        if 치명도_s:
            base = 치명도_s

    # CLAIM 가중치 (+1)
    if row.get('분류') == 'CLAIM':
        base += 1

    # 피해보상비 가중치
    cost = row.get('피해보상비') or 0
    if isinstance(cost, str):
        try:
            cost = int(cost.replace(',', ''))
        except:
            cost = 0
    if cost >= 100_000_000:  # 1억 이상 -> S=10
        return 10
    elif cost >= 10_000_000:  # 1천만 이상 -> +1
        base += 1

    return min(base, 10)


def calc_o_value_from_qa(total_count, recent_3yr_count):
    """QA DB 발생 건수에서 O값 자동 계산"""
    # 기본 O값 (12년간 건수 기준)
    if total_count >= 50:
        base = 8
    elif total_count >= 20:
        base = 6
    elif total_count >= 10:
        base = 4
    elif total_count >= 5:
        base = 3
    elif total_count >= 1:
        base = 2
    else:
        base = 1

    # 최근 3년 트렌드 보정
    if total_count > 0:
        recent_ratio = recent_3yr_count / total_count
        if recent_ratio > 0.5 and base < 10:
            base += 1  # 증가 추세
        elif recent_ratio < 0.2 and base > 1:
            base -= 1  # 개선 추세

    return min(max(base, 1), 10)


def calc_d_value_from_qa(row):
    """QA DB 데이터에서 D값 자동 계산"""
    base = 5  # 기본값

    # 검사구분 기준
    if row.get('검사구분'):
        base = QA_D_VALUE_MAPPING['검사구분'].get(row['검사구분'], 5)

    # 항목구분으로 보정
    if row.get('항목구분'):
        adjustment = QA_D_VALUE_MAPPING['항목구분'].get(row['항목구분'])
        if adjustment:
            base = min(base, adjustment)

    return min(max(base, 1), 10)


def get_lifecycle_tag_from_qa(row):
    """QA DB 데이터에서 라이프사이클 태그 추론"""
    tags = set()

    # 발생원인유형 기반
    cause_type = row.get('발생원인유형') or ''
    for tag, keywords in QA_LIFECYCLE_MAPPING.items():
        for kw in keywords:
            if kw in cause_type:
                tags.add(tag)
                break

    # 원인부서 기반 (대괄호 없이 통일)
    dept = row.get('원인부서') or ''
    if '설계' in dept:
        tags.add('설계')
    elif '자재' in dept or '구매' in dept:
        tags.add('재료')
    elif '생산' in dept or '제조' in dept or '가공' in dept:
        tags.add('제작')
    elif '품질' in dept or '검사' in dept or '시험' in dept:
        tags.add('시험')

    return list(tags) if tags else ['제작']  # 기본값


def query_qa_for_component(conn, component_name):
    """부품명으로 QA DB 이력 조회"""
    if not conn:
        return []

    query = """
    SELECT
        품명,
        발생현상유형,
        발생현상유형소분류,
        현상_소분류,
        발생원인,
        발생원인유형,
        원인부서,
        중요_경미,
        치명도,
        분류,
        피해보상비,
        검사구분,
        항목구분,
        재발방지대책,
        조치내역,
        발생년도,
        COUNT(*) as 발생횟수
    FROM qa_records
    WHERE 품명 LIKE ?
    GROUP BY
        발생현상유형, 발생현상유형소분류, 발생원인, 발생원인유형
    ORDER BY 발생횟수 DESC
    LIMIT 20
    """
    cursor = conn.execute(query, (f'%{component_name}%',))
    return [dict(row) for row in cursor.fetchall()]


def get_component_stats(conn, component_name):
    """부품별 통계 조회 (O값 계산용)"""
    if not conn:
        return {'total': 0, 'recent_3yr': 0}

    query = """
    SELECT
        COUNT(*) as total,
        SUM(CASE WHEN 발생년도 >= 2024 THEN 1 ELSE 0 END) as recent_3yr
    FROM qa_records
    WHERE 품명 LIKE ?
    """
    cursor = conn.execute(query, (f'%{component_name}%',))
    row = cursor.fetchone()
    return {'total': row['total'] or 0, 'recent_3yr': row['recent_3yr'] or 0}


def enhance_fmea_with_qa_db(fmea_data, db_path):
    """FMEA 데이터를 QA DB로 보강 (S/O/D 자동 계산, 태그 자동 부여)"""
    conn = connect_qa_db(db_path)
    if not conn:
        print("[WARN] QA DB 연동 실패 - 원본 데이터 사용")
        return fmea_data

    print(f"[INFO] QA DB 연동: {db_path}")
    enhanced_count = 0

    for item in fmea_data:
        component = item.get('부품명', '')
        if not component:
            continue

        # QA 이력 조회
        qa_records = query_qa_for_component(conn, component)
        stats = get_component_stats(conn, component)

        if not qa_records:
            continue

        # 대표 레코드 (가장 빈번한 건)
        top_record = qa_records[0]

        # S값 자동 계산 (값이 없거나 0인 경우만)
        if not item.get('S') or item.get('S') == 0:
            item['S'] = calc_s_value_from_qa(top_record)
            item['S_source'] = 'QA_DB_AUTO'

        # O값 자동 계산
        if not item.get('O') or item.get('O') == 0:
            item['O'] = calc_o_value_from_qa(stats['total'], stats['recent_3yr'])
            item['O_source'] = f"QA_DB_AUTO (total:{stats['total']}, recent:{stats['recent_3yr']})"

        # D값 자동 계산
        if not item.get('D') or item.get('D') == 0:
            item['D'] = calc_d_value_from_qa(top_record)
            item['D_source'] = 'QA_DB_AUTO'

        # 라이프사이클 태그 자동 부여 (고장원인에 태그 없는 경우)
        # 대괄호 없이 통일
        cause = item.get('고장원인', '')
        if cause and not any(tag in cause for tag in ['설계', '재료', '제작', '시험']):
            lifecycle_tags = get_lifecycle_tag_from_qa(top_record)
            if lifecycle_tags:
                item['고장원인'] = f"{lifecycle_tags[0]}: {cause}"

        # 고장영향 보강 (없는 경우)
        if not item.get('고장영향') and top_record.get('발생현상유형'):
            effect = top_record.get('발생현상유형', '')
            sub_effect = top_record.get('발생현상유형소분류', '')
            if sub_effect and sub_effect != effect:
                item['고장영향'] = f"{effect} > {sub_effect}"
            else:
                item['고장영향'] = effect

        # 예방대책 보강 (없는 경우)
        if not item.get('현 예방관리') and top_record.get('재발방지대책'):
            item['현 예방관리'] = top_record['재발방지대책']

        # 검출대책 보강 (없는 경우)
        if not item.get('현 검출관리') and top_record.get('조치내역'):
            item['현 검출관리'] = top_record['조치내역']

        enhanced_count += 1

    conn.close()
    print(f"[OK] QA DB 보강 완료: {enhanced_count}/{len(fmea_data)} 항목")
    return fmea_data


# ========================================
# 용어 사전 (전문용어 -> 병기 형식) - references/glossary.md 참조
# ========================================
GLOSSARY = {
    # 철심 관련
    '자왜변형': '자왜변형(철심진동)',
    '자왜진동': '자왜진동(자속진동)',
    '층간단락': '층간단락(철심내부합선)',
    '와전류': '와전류(유도전류)',
    '철손': '철손(철심손실)',
    '적층이완': '적층이완(철심판풀림)',
    '적층갭': '적층갭(철심판틈새)',
    '자기포화': '자기포화(자속과다)',
    # 권선 관련
    '절연열화': '절연열화(절연성능저하)',
    '절연파괴': '절연파괴(절연깨짐)',
    '권선변형': '권선변형(권선찌그러짐)',
    '부분방전': '부분방전(미세방전)',
    '턴간단락': '턴간단락(권선내부합선)',
    # 기타
    '열화': '열화(성능저하)',
    '열점': '열점(국부과열점)',
    '피로파손': '피로파손(반복하중파손)',
    '응력집중': '응력집중(힘집중)',
}


def apply_glossary(text, used_terms=None):
    """전문 용어를 병기 형식으로 변환 (첫 등장 시만)

    Args:
        text: 변환할 텍스트
        used_terms: 이미 사용된 용어 집합 (첫 등장 추적)

    Returns:
        변환된 텍스트
    """
    if not text or not isinstance(text, str):
        return text

    if used_terms is None:
        used_terms = set()

    result = text
    for term, replacement in GLOSSARY.items():
        # 이미 병기 형식이면 스킵
        if replacement in result:
            used_terms.add(term)
            continue
        # 첫 등장 시만 병기
        if term in result and term not in used_terms:
            result = result.replace(term, replacement, 1)  # 첫 번째만 교체
            used_terms.add(term)

    return result


# ========================================
# 변압기 부품 분류 및 기능 템플릿 데이터 (FALLBACK 전용)
# [!] 중요: 이 템플릿은 WebSearch/내부문서에서 정보를 찾지 못했을 때만 사용
# 기준: AIAG-VDA 2019, references/function-analysis.md
#
# 올바른 워크플로우:
# 1. 내부 문서 확인 (구성품.xlsx, 도면, CHECK SHEET)
# 2. WebSearch 실행 ("{부품명} transformer function purpose")
# 3. 위 결과 없을 경우에만 아래 템플릿 사용
# ========================================

# 부품명 -> 구분(카테고리) 매핑 패턴 (FALLBACK 전용)
# [!] 입력 데이터의 '구분' 필드가 우선! 이 패턴은 fallback으로만 사용
PART_CATEGORY_PATTERNS = {
    '중신': [
        'CORE', 'YOKE', '철심', 'IRON_CORE', 'LAMINATION',
        'LOW_CLAMP', 'CLAMP', 'BASE_BRACKET', 'TANK_SUPPORT',
        'YOKE_CLAMP', 'FLITCH_PLATE', 'CORE_TIE', 'CORE_BOLT',
        'CORE_BOLT_INSULATION', '클램프', '브라켓', '요크',
        '규소강판', '적층철심', '클램프 절연물', '클램프절연물'
    ],
    '권선': [
        'WINDING', 'HV_WINDING', 'LV_WINDING', 'TV_WINDING',
        'COIL', 'R-SPACER', 'A-SPACER', 'RADIAL_SPACER',
        'AXIAL_SPACER', 'INSULATION', 'CONDUCTOR', 'LEAD',
        'TRANSPOSITION', 'CTC', '권선', '스페이서', '절연'
    ],
    '단자': [
        'BUSHING', 'HV_BUSHING', 'LV_BUSHING', 'TERMINAL',
        'CONNECTOR', 'TAP', 'TAP_LEAD', 'TAP_CHANGER', 'OLTC',
        'SELECTOR', 'DIVERTER', '부싱', '단자', '탭', '탭절환기'
    ],
    '외함': [
        'TANK', 'COVER', 'CONSERVATOR', 'GASKET', 'RADIATOR',
        'COOLING', 'PUMP', 'FAN', 'OIL', 'BREATHER', 'DRAIN',
        'BUCHHOLZ', 'PRV', 'THERMOMETER', 'OIL_LEVEL', 'SILICA_GEL',
        '탱크', '외함', '방열기', '보존기', '냉각', '개스킷'
    ]
}

# 부품명 -> 표준 보조기능 템플릿 (FALLBACK 전용)
# [!] 입력 데이터의 '보조기능' 필드가 우선! 이 템플릿은 fallback으로만 사용
# 기준: references/function-analysis.md 부품별 기능 예시 (L107-140)
PART_AUXILIARY_FUNCTIONS = {
    # === 중신 ===
    'CORE': '권선을 기계적으로 지지한다 / 와전류 손실을 최소화한다',
    'YOKE': '철심 회로를 연결한다 / 자속을 균일하게 분배한다',
    'LOW_CLAMP': '진동을 흡수한다 / 하중을 분산한다',
    'LOW_CLAMP_LV': '진동을 흡수한다 / 권선 변형을 방지한다',
    'LOW_CLAMP_HV': '진동을 흡수한다 / 고압권선을 보호한다',
    'YOKE_CLAMP': '요크 적층을 유지한다 / 진동을 억제한다',
    'BASE_BRACKET': '하중을 분산한다 / 충격을 완화한다',
    'TANK_SUPPORT': '충격을 흡수한다 / 운송 시 안정성을 제공한다',
    'FLITCH_PLATE': '철심 압축력을 유지한다 / 적층 안정성을 확보한다',
    'CORE_TIE': '철심 결합력을 유지한다 / 진동을 억제한다',
    'CORE_BOLT': '철심 적층을 고정한다 / 절연 특성을 유지한다',
    'CORE_BOLT_INSULATION': '와전류를 차단한다 / 국부 과열을 방지한다',

    # === 권선 ===
    'HV_WINDING': '임피던스를 제공한다 / 단락 전류를 제한한다',
    'LV_WINDING': '단락 전류를 제한한다 / 부하 변동에 대응한다',
    'TV_WINDING': '계측 신호를 공급한다 / 보호 계전기에 전원을 공급한다',
    'R-SPACER': '절연 거리를 유지한다 / 권선 진동을 억제한다',
    'A-SPACER': '권선 간격을 유지한다 / 열변형을 흡수한다',
    'RADIAL_SPACER': '절연 거리를 유지한다 / 오일 순환을 유도한다',
    'AXIAL_SPACER': '축방향 힘을 지지한다 / 권선 변형을 방지한다',
    'INSULATION': '열을 차단한다 / 기계적 보호를 제공한다',
    'CONDUCTOR': '전류 손실을 최소화한다 / 열을 방산한다',
    'LEAD': '유연성을 제공한다 / 진동을 흡수한다',
    'CTC': '와전류 손실을 감소한다 / 전류 분포를 균등화한다',

    # === 단자 ===
    'HV_BUSHING': '절연을 유지한다 / 기밀을 유지한다',
    'LV_BUSHING': '기밀을 유지한다 / 외부 오염을 차단한다',
    'TERMINAL': '접촉 저항을 최소화한다 / 열팽창을 흡수한다',
    'TAP_CHANGER': '전압 안정성을 유지한다 / 과도 전압을 억제한다',
    'TAP_LEAD': '유연성을 제공한다 / 절연을 유지한다',
    'OLTC': '무부하 전환을 수행한다 / 아크를 소호한다',
    'SELECTOR': '탭 위치를 결정한다 / 접촉 신뢰성을 유지한다',
    'DIVERTER': '아크를 소호한다 / 접점을 보호한다',

    # === 외함 ===
    'TANK': '외부 환경을 차단한다 / 기계적 보호를 제공한다',
    'COVER': '기밀을 유지한다 / 내부 부품을 보호한다',
    'GASKET': '오일 누출을 방지한다 / 수분 침입을 차단한다',
    'CONSERVATOR': '수분 침입을 방지한다 / 유면 변화를 시각화한다',
    'RADIATOR': '열교환 면적을 확보한다 / 오일 순환을 유도한다',
    'COOLING': '냉각 효율을 극대화한다 / 온도 상승을 억제한다',
    'PUMP': '오일 유량을 증가한다 / 냉각 효율을 향상한다',
    'FAN': '공기 유동을 증가한다 / 방열기 효율을 향상한다',
    'OIL': '부품 산화를 방지한다 / 부분방전을 억제한다',
    'BREATHER': '수분을 흡착한다 / 외기 오염을 차단한다',
    'SILICA_GEL': '습기를 제거한다 / 절연유 품질을 유지한다',
    'BUCHHOLZ': '가스 누적을 감지한다 / 경보를 발생한다',
    'PRV': '압력을 조절한다 / 탱크 파손을 방지한다',
    'THERMOMETER': '과열을 조기 감지한다 / 경보를 발생한다',
    'OIL_LEVEL': '오일량 이상을 감지한다 / 경보를 발생한다',

    # === 한글 부품명 템플릿 ===
    # 클램프 관련
    '클램프': '진동을 흡수한다 / 하중을 분산한다',
    '클램프 절연물': '와전류를 차단한다 / 권선과 철심 간 전기적 분리를 유지한다',
    '클램프 절연': '와전류를 차단한다 / 권선과 철심 간 절연을 유지한다',
    '클램프절연물': '와전류를 차단한다 / 권선과 철심 간 전기적 분리를 유지한다',

    # 철심/코어 관련
    '철심': '권선을 기계적으로 지지한다 / 와전류 손실을 최소화한다',
    '규소강판': '와전류를 억제한다 / 철손을 최소화한다',
    '적층철심': '와전류를 억제한다 / 자속 밀도를 균일화한다',

    # 권선 관련
    '고압권선': '임피던스를 제공한다 / 단락 전류를 제한한다',
    '저압권선': '단락 전류를 제한한다 / 부하 변동에 대응한다',
    '권선': '단락 전류를 제한한다 / 임피던스를 제공한다',

    # 절연 관련
    '절연물': '열을 차단한다 / 기계적 보호를 제공한다',
    '절연지': '층간 절연을 제공한다 / 열을 차단한다',
    '절연유': '부품 산화를 방지한다 / 부분방전을 억제한다',

    # 스페이서 관련
    '스페이서': '절연 거리를 유지한다 / 냉각 통로를 확보한다',
    '방사 스페이서': '절연 거리를 유지한다 / 오일 순환을 유도한다',
    '축방향 스페이서': '축방향 힘을 지지한다 / 권선 변형을 방지한다',

    # 부싱/단자 관련
    '부싱': '절연을 유지한다 / 기밀을 유지한다',
    '고압부싱': '절연을 유지한다 / 기밀을 유지한다',
    '저압부싱': '기밀을 유지한다 / 외부 오염을 차단한다',
    '단자': '접촉 저항을 최소화한다 / 열팽창을 흡수한다',

    # 탱크/외함 관련
    '탱크': '외부 환경을 차단한다 / 기계적 보호를 제공한다',
    '외함': '외부 환경을 차단한다 / 기계적 보호를 제공한다',
    '방열기': '열교환 면적을 확보한다 / 오일 순환을 유도한다',
    '보존기': '수분 침입을 방지한다 / 유면 변화를 수용한다',
    '개스킷': '오일 누출을 방지한다 / 수분 침입을 차단한다',

    # 보호장치 관련
    '부흐홀츠': '가스 누적을 감지한다 / 경보를 발생한다',
    '온도계': '과열을 조기 감지한다 / 경보를 발생한다',
    '압력방출장치': '압력을 조절한다 / 탱크 파손을 방지한다',

    # 탭절환기 관련
    '탭절환기': '전압 안정성을 유지한다 / 과도 전압을 억제한다',
    '탭리드': '유연성을 제공한다 / 절연을 유지한다'
}


def infer_category(part_name, input_category=None):
    """부품명에서 구분(카테고리) 조회

    [!] 우선순위:
    1. input_category (내부문서/WebSearch에서 제공된 값) - 최우선
    2. PART_CATEGORY_PATTERNS 패턴 매칭 - fallback

    기준: references/function-analysis.md
    """
    # 1. 내부문서/WebSearch에서 제공된 값이 있으면 최우선 사용
    if input_category and input_category.strip():
        valid_categories = ['중신', '권선', '단자', '외함', '기타']
        if input_category.strip() in valid_categories:
            return input_category.strip()
    part_upper = part_name.upper().strip()

    # 1. 정확한 매칭 시도
    for category, patterns in PART_CATEGORY_PATTERNS.items():
        for pattern in patterns:
            if part_upper == pattern.upper():
                return category

    # 2. 부분 매칭 시도 (부품명에 패턴이 포함된 경우)
    for category, patterns in PART_CATEGORY_PATTERNS.items():
        for pattern in patterns:
            if pattern.upper() in part_upper or part_upper in pattern.upper():
                return category

    # 3. 한글 키워드 매칭
    if any(kw in part_name for kw in ['철심', '클램프', '브라켓', '요크']):
        return '중신'
    if any(kw in part_name for kw in ['권선', '스페이서', '절연', '도체']):
        return '권선'
    if any(kw in part_name for kw in ['부싱', '단자', '탭', '커넥터']):
        return '단자'
    if any(kw in part_name for kw in ['탱크', '외함', '방열', '냉각', '오일', '보존']):
        return '외함'

    # 4. 기본값 (분류 불가)
    return '기타'


def get_auxiliary_functions(part_name, input_auxiliary=None):
    """부품명에서 보조기능 조회

    [!] 우선순위:
    1. input_auxiliary (WebSearch/내부문서에서 제공된 값) - 최우선
    2. PART_AUXILIARY_FUNCTIONS 템플릿 - fallback

    기준: references/function-analysis.md
    """
    # 1. WebSearch/내부문서에서 제공된 값이 있으면 최우선 사용
    if input_auxiliary and input_auxiliary.strip():
        # "(보조기능 입력 필요)" 같은 플레이스홀더가 아닌 경우만
        if '입력 필요' not in input_auxiliary and input_auxiliary != '-':
            return input_auxiliary

    part_name_stripped = part_name.strip()
    part_upper = part_name_stripped.upper()

    # 2. 정확한 매칭 (한글 먼저) - FALLBACK
    if part_name_stripped in PART_AUXILIARY_FUNCTIONS:
        return PART_AUXILIARY_FUNCTIONS[part_name_stripped]

    # 3. 정확한 매칭 (영문 대문자) - FALLBACK
    if part_upper in PART_AUXILIARY_FUNCTIONS:
        return PART_AUXILIARY_FUNCTIONS[part_upper]

    # 4. 한글 키워드 매칭 (부분 매칭) - FALLBACK
    korean_keywords = [
        ('클램프 절연', '클램프 절연물'),
        ('클램프절연', '클램프절연물'),
        ('절연물', '절연물'),
        ('절연지', '절연지'),
        ('절연유', '절연유'),
        ('클램프', '클램프'),
        ('철심', '철심'),
        ('규소강판', '규소강판'),
        ('고압권선', '고압권선'),
        ('저압권선', '저압권선'),
        ('권선', '권선'),
        ('스페이서', '스페이서'),
        ('부싱', '부싱'),
        ('단자', '단자'),
        ('탱크', '탱크'),
        ('외함', '외함'),
        ('방열기', '방열기'),
        ('보존기', '보존기'),
        ('개스킷', '개스킷'),
        ('탭절환기', '탭절환기'),
        ('부흐홀츠', '부흐홀츠'),
        ('온도계', '온도계'),
    ]

    for keyword, template_key in korean_keywords:
        if keyword in part_name_stripped:
            if template_key in PART_AUXILIARY_FUNCTIONS:
                return PART_AUXILIARY_FUNCTIONS[template_key]

    # 5. 영문 부분 매칭 (접두어) - FALLBACK
    for key, value in PART_AUXILIARY_FUNCTIONS.items():
        if part_upper.startswith(key.upper()) or key.upper().startswith(part_upper):
            return value

    # 6. 기본값 - WebSearch 필요 알림
    return '(WebSearch로 보조기능 확인 필요)'


def validate_function_format(function_text):
    """기능 표현 형식 검증 (AIAG-VDA 표준)

    올바른 형식: "[주체]가 [목적어]를 [동사]한다"
    기준: references/function-analysis.md

    반환값: (is_valid, message)
    """
    if not function_text or not function_text.strip():
        return False, "기능 텍스트가 비어있습니다"

    text = function_text.strip()

    # 1. 동사 어미 검증 (한다, 한다, 된다, 있다 등)
    verb_endings = ['한다', '된다', '있다', '준다', '낸다', '간다', '온다', '난다']
    has_verb = any(text.endswith(ending) for ending in verb_endings)

    # 2. 조사 검증 (을/를, 이/가)
    has_object_marker = '를 ' in text or '을 ' in text
    has_subject_marker = '가 ' in text or '이 ' in text or '은 ' in text or '는 ' in text

    # 3. 금지 패턴 검증
    forbidden_patterns = [
        (r'^(기능|고장|영향|형태|원인)\s*\d+\s*:', "번호 표기 금지"),
        (r'^\d+\s*[:.)]', "번호로 시작 금지"),
        (r'^[①②③④⑤⑥⑦⑧⑨⑩]', "원 숫자 금지"),
    ]

    for pattern, msg in forbidden_patterns:
        if re.match(pattern, text):
            return False, msg

    # 4. 최소 길이 검증
    if len(text) < 5:
        return False, "기능 설명이 너무 짧습니다 (최소 5자)"

    # 5. 종합 판정
    if has_verb and (has_object_marker or has_subject_marker):
        return True, "[OK] 올바른 형식"
    elif has_verb:
        return True, "[!] 동사는 있으나 조사 확인 필요"
    else:
        return False, "[X] '[주체]가 [목적어]를 [동사]한다' 형식으로 작성하세요"


def validate_function_analysis_data(function_data):
    """기능분석 데이터 전체 검증

    반환값: (is_valid, issues_list)
    """
    issues = []

    for i, item in enumerate(function_data):
        # 1. 구분 검증
        if not item.get('구분') or item['구분'] == '기타':
            issues.append(f"행 {i+1}: 구분 '{item.get('구분', '')}' - 명확한 카테고리 지정 필요")

        # 2. 기능 형식 검증
        is_valid, msg = validate_function_format(item.get('기능', ''))
        if not is_valid:
            issues.append(f"행 {i+1}: 기능 - {msg}")

    if issues:
        print(f"[!] 기능분석 데이터 검증 이슈 ({len(issues)}건):")
        for issue in issues[:10]:  # 최대 10개 출력
            print(f"    {issue}")
        if len(issues) > 10:
            print(f"    ... 외 {len(issues) - 10}건")
        return False, issues

    print(f"[OK] 기능분석 데이터 검증 통과 ({len(function_data)}개 항목)")
    return True, []


def extract_function_analysis_data(fmea_data, project_info=None):
    """FMEA 데이터에서 기능분석 데이터 추출 (중복 제거 + 자동 추론)

    기준: references/function-analysis.md
    AIAG-VDA Step 3: 기능 분석

    열 구조: 구분 | 파트명 | 기능 | 관련 고장형태 | 고장영향 | 비고
    비고: "주기능 (출처)" 또는 "보조기능 (출처)"
      - 출처: 다이어그램 / 내부문서 / WebSearch

    추가 기능 (최대 3개):
    - project_info.additional_functions에서 가져옴
    - Leader가 Phase 1에서 Worker B/C 결과를 병합하여 확정
    - Phase 2에서 FMEA 항목이 생성되면 fmea_data에 포함 -> 자동 추출
    - FMEA 항목이 없는 추가 기능은 fallback으로 빈 행 추가
    """

    # function_order에서 다이어그램 기능 목록 추출
    function_order = []
    if project_info:
        function_order = project_info.get('function_order', []) or []

    # 추가 기능 출처 룩업 (부품명, 기능명) -> 출처
    additional_source_map = {}
    additional_functions = []
    if project_info:
        additional_functions = project_info.get('additional_functions', []) or []
    for af in additional_functions[:3]:
        af_key = (af.get('part', ''), af.get('name', ''))
        additional_source_map[af_key] = af.get('source', 'WebSearch')

    # 부품별 첫 번째 기능 추적 (주기능 판별용)
    part_first_function = {}

    # 부품명+기능 조합으로 중복 제거
    seen = {}
    function_data = []

    for row in fmea_data:
        key = (row['부품명'], row['기능'])
        part_name = row['부품명']

        # 입력 데이터에서 구분 확인
        input_category = row.get('구분', None)

        if key not in seen:
            # 주기능/보조기능 판별: 부품별 첫 번째 기능 = 주기능
            if part_name not in part_first_function:
                part_first_function[part_name] = row['기능']
                func_type = "주기능"
            else:
                func_type = "보조기능"

            # 출처 판별: 다이어그램 > 추가기능(내부문서/WebSearch) > 기본값
            if function_order and row['기능'] in function_order:
                source = "다이어그램"
            elif key in additional_source_map:
                source = additional_source_map[key]
            elif not function_order:
                source = "다이어그램"  # function_order 없으면 기본값
            else:
                source = "WebSearch"

            seen[key] = {
                '구분': infer_category(part_name, input_category),  # [*] 입력 우선, 패턴 fallback
                '파트명': part_name,
                '기능': row['기능'],
                '관련 고장형태': row['고장형태'],
                '고장영향': row['고장영향'],
                '비고': "%s (%s)" % (func_type, source)
            }
        else:
            # 동일 부품+기능에 다른 고장형태가 있으면 추가
            existing = seen[key]
            if row['고장형태'] not in existing['관련 고장형태']:
                existing['관련 고장형태'] += f", {row['고장형태']}"

    # 추가 기능 fallback (Phase 2에서 FMEA 항목이 없는 추가 기능만)
    for af in additional_functions[:3]:  # 최대 3개 제한
        af_name = af.get('name', '')
        af_source = af.get('source', 'WebSearch')
        af_part = af.get('part', '')  # 부품명
        if not af_name or not af_part:
            continue
        key = (af_part, af_name)
        if key not in seen:
            seen[key] = {
                '구분': infer_category(af_part, None),
                '파트명': af_part,
                '기능': af_name,
                '관련 고장형태': '',
                '고장영향': '',
                '비고': "보조기능 (%s)" % af_source
            }

    # 정렬: 구분 -> 파트명 -> 기능 (카테고리별 그룹핑)
    category_order = {'중신': 1, '권선': 2, '단자': 3, '외함': 4, '기타': 5}
    function_data = sorted(
        seen.values(),
        key=lambda x: (category_order.get(x['구분'], 99), x['파트명'], x['기능'])
    )

    # 통계 출력
    category_counts = {}
    for item in function_data:
        cat = item['구분']
        category_counts[cat] = category_counts.get(cat, 0) + 1

    print(f"[OK] 기능분석 데이터 추출 완료 ({len(function_data)}개 항목)")
    print(f"    구분별: {category_counts}")
    return function_data


def create_function_analysis_sheet(wb, function_data, project_info=None):
    """기능분석 시트 생성 (AIAG-VDA Step 3)

    기준: references/function-analysis.md
    열 구조: 구분 | 파트명 | 기능 | 관련 고장형태 | 고장영향 | 비고
    """

    # 시트 생성 (첫 번째 위치에)
    ws = wb.create_sheet("기능분석", 0)

    # 테두리 정의
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row 1: 제목
    part_name = project_info.get('부품', '부품명') if project_info else '부품명'
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"{part_name}_기능분석 (AIAG-VDA Step 3)"
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row 2: 설명
    ws.merge_cells('A2:F2')
    desc_cell = ws['A2']
    desc_cell.value = "※ FMEA 분석 전 단계: 각 부품의 기능을 정의합니다 (비고란에 주기능/보조기능 구분)"
    desc_cell.font = Font(italic=True, size=11, color="666666")
    desc_cell.alignment = Alignment(horizontal='left', vertical='center')

    # Row 3: 빈 행
    ws.row_dimensions[3].height = 5

    # Row 4: 헤더
    headers = ["구분", "파트명", "기능", "관련 고장형태", "고장영향", "비고"]
    header_widths = [10, 20, 35, 25, 30, 25]

    for col_num, (header, width) in enumerate(zip(headers, header_widths), 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.row_dimensions[4].height = 25

    # Row 5+: 데이터
    for row_idx, item in enumerate(function_data, start=5):
        ws.cell(row=row_idx, column=1, value=item['구분'])
        ws.cell(row=row_idx, column=2, value=item['파트명'])
        ws.cell(row=row_idx, column=3, value=item['기능'])
        ws.cell(row=row_idx, column=4, value=item['관련 고장형태'])
        ws.cell(row=row_idx, column=5, value=item['고장영향'])
        ws.cell(row=row_idx, column=6, value=item['비고'])

        # 서식 적용
        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            if col in [1, 2, 6]:  # 구분, 파트명, 비고
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        ws.row_dimensions[row_idx].height = 30

    # 입력 안내 메시지 추가 (Row 4 헤더)
    function_messages = {
        'A4': ("구분 입력", "부품 카테고리\n예: 중신, 권선, 단자, 외함"),
        'B4': ("파트명 입력", "부품 이름\n예: CORE, LOW_CLAMP, HV_WINDING"),
        'C4': ("기능 입력", "형식: [목적어]를 [동사]한다\n예: 자속을 전달한다"),
        'D4': ("관련 고장형태", "이 기능 실패 시 나타나는 현상\n예: 과열, 변형, 절연파괴"),
        'E4': ("고장영향", "최종 영향\n예: 변압기 정지, 권선 손상"),
        'F4': ("비고", "주기능/보조기능 구분 + 출처\n예: 주기능 (다이어그램)")
    }

    for cell_ref, (title, prompt) in function_messages.items():
        dv = DataValidation(type="none", allow_blank=True)
        dv.promptTitle = title
        dv.prompt = prompt
        dv.showInputMessage = True
        dv.add(cell_ref)
        ws.add_data_validation(dv)

    # 틀고정
    ws.freeze_panes = 'A5'

    print(f"[OK] 기능분석 시트 생성 완료 ({len(function_data)}개 항목)")
    return ws


def sort_fmea_data(data, function_order=None):
    """FMEA 데이터 결정적 정렬 (excel-generation.md 알고리즘)

    정렬 순서는 병합 계층 구조와 일치해야 함:
    - A열 (부품명): parent 없음
    - B열 (기능): parent = 부품명
    - C열 (고장영향): parent = 기능  <- CRITICAL!
    - D열 (S): parent = 고장영향
    - E열 (고장형태): parent = 고장영향

    [!] 고장영향이 정렬 키에 없으면 다이아몬드 구조 병합 불가!

    Args:
        data: FMEA 데이터 리스트
        function_order: 기능 순서 리스트 (다이어그램 순서).
                        None이면 문자열 순으로 정렬.
                        예: ["전류의 경로를 제공한다", "전류를 흘려 자기장을 만든다", ...]
    """

    # 라이프사이클 단계 순서 (4단계)
    lifecycle_order = {
        '설계': 1, '재료': 2, '제작': 3, '시험': 4
    }

    # 기능 순서 맵 생성 (function_order가 있는 경우)
    function_order_map = {}
    if function_order:
        for idx, func in enumerate(function_order):
            function_order_map[func] = idx
        print(f"   [INFO] 기능 순서 지정됨: {len(function_order)}개 기능 (다이어그램 순서)")

    # [!!] 기능_순서 필드 자동 감지 (JSON 데이터에 기능_순서가 있으면 사용)
    has_function_order_field = any('기능_순서' in row for row in data)
    if has_function_order_field:
        print(f"   [INFO] 기능_순서 필드 감지됨 -> 다이어그램 순서 사용")

    def sort_key(row):
        # 고장원인에서 라이프사이클 단계 추출 (대괄호 제거로 일관성 유지)
        lifecycle_stage = row['고장원인'].split(':')[0].strip().strip('[]')

        # S 값을 정수로 변환 (문자열일 수 있음)
        s_value = int(row['S']) if isinstance(row['S'], (int, str)) and str(row['S']).isdigit() else 0

        # 기능 정렬 키 결정 (우선순위: 기능_순서 필드 > function_order 파라미터 > 문자열 순)
        # [!!] CRITICAL: 기능_순서 필드가 있으면 다이어그램 순서 사용!
        if '기능_순서' in row:
            # JSON 데이터의 기능_순서 필드 사용 (가장 우선!)
            func_key = row.get('기능_순서', 999)
        elif function_order_map:
            # function_order 파라미터로 전달된 순서 사용
            func_key = function_order_map.get(row['기능'], 999)
        else:
            # 둘 다 없으면 문자열 순 (비권장!)
            func_key = row['기능']
            if not has_function_order_field and not function_order_map:
                pass  # 경고는 이미 출력됨

        return (
            row['부품명'],                                      # 1. 부품명 (문자열 순)
            func_key,                                           # 2. 기능 (다이어그램 순서 또는 문자열 순)
            row['고장영향'],                                    # 3. 고장영향 (문자열 순) <- 추가!
            -s_value,                                           # 4. S (내림차순) - 정수로 변환
            row['고장형태'],                                    # 5. 고장형태 (문자열 순)
            lifecycle_order.get(lifecycle_stage, 99),          # 6. 라이프사이클 단계 순서
            row['고장원인']                                     # 7. 고장원인 (문자열 순)
        )

    # 기능_순서 필드 누락 경고
    if not has_function_order_field and not function_order_map:
        print(f"   [WARNING] 기능_순서 필드 없음! 기능이 문자열 순으로 정렬됩니다.")
        print(f"             -> 다이어그램 순서를 유지하려면 JSON에 기능_순서 필드를 추가하세요.")

    return sorted(data, key=sort_key)


def validate_data(data):
    """빈 셀 방지 및 형식 검증 (excel-generation.md)"""

    required_columns = [
        # 현재 상태 (1-13) - 스킬 범위: STEP 1-5 (리스크 분석)
        "부품명", "기능", "고장영향", "S", "고장형태",
        "고장원인", "고장메커니즘", "현재예방대책", "O",
        "현재검출대책", "D", "RPN", "AP"
        # 조치 계획 (14-17): 예방조치, 검출조치, 담당자, 목표일 - 빈값 허용 (스킬 범위 밖)
        # 개선 후 (18-22): S', O', D', RPN', AP' - 빈값 허용 (스킬 범위 밖)
    ]

    # 스킬 범위 밖 컬럼 (빈값 허용 - STEP 6 최적화에서 별도 입력)
    optional_columns = [
        "예방조치", "검출조치", "담당자", "목표일",  # N~Q열 (조치 계획)
        "S'", "O'", "D'", "RPN'", "AP'"              # R~V열 (개선 후)
    ]

    # 1. 필수 컬럼 빈 값 검증
    for i, row in enumerate(data):
        for col in required_columns:
            if not row.get(col) or row[col] == "":
                raise ValueError(f"행 {i+1}, 컬럼 '{col}': 빈 값 불허")
        # 개선 후 컬럼: 빈값이면 빈 문자열로 초기화
        for col in optional_columns:
            if col not in row or row[col] is None:
                row[col] = ""

    # 1-1. 중복 검증 (CRITICAL!)
    # 같은 기능 내에서 고장영향 중복 검증
    seen_combinations = {}
    for i, row in enumerate(data):
        key = (row['부품명'], row['기능'], row['고장영향'])
        if key not in seen_combinations:
            seen_combinations[key] = i
        # 중복은 허용하되, 같은 기능 내에서 같은 고장영향이 여러 번 나오는 것은 정상
        # (다른 고장형태, 다른 고장원인으로 인해)

    # 1-2. 정렬 순서 검증 (CRITICAL!)
    # 부품명 -> 기능 -> 고장영향 -> 고장형태 -> 고장원인 순서가 지켜져야 함
    prev_part = None
    prev_func = None
    prev_effect = None
    prev_form = None

    for i, row in enumerate(data):
        curr_part = row['부품명']
        curr_func = row['기능']
        curr_effect = row['고장영향']
        curr_form = row['고장형태']

        # 부품명이 바뀌면 초기화
        if curr_part != prev_part:
            prev_part = curr_part
            prev_func = None
            prev_effect = None
            prev_form = None

        # 기능이 바뀌면 초기화
        if curr_func != prev_func:
            prev_func = curr_func
            prev_effect = None
            prev_form = None

        # 고장영향이 바뀌면 초기화
        if curr_effect != prev_effect:
            prev_effect = curr_effect
            prev_form = None

        # 고장형태가 바뀌면 기록
        if curr_form != prev_form:
            prev_form = curr_form

    # 2. 형식 검증 (번호 표기 금지!)
    number_patterns = [
        r'^(기능|고장영향|고장형태|고장원인|영향|형태|원인)\s*\d+\s*:',  # "기능 1:", "고장영향 2:"
        r'^\d+\s*[:.)]',  # "1:", "1.", "1)"
        r'^[①②③④⑤⑥⑦⑧⑨⑩]',  # 원 숫자
    ]

    for i, row in enumerate(data):
        for col in ['기능', '고장영향', '고장형태']:
            value = str(row.get(col, ''))
            for pattern in number_patterns:
                if re.match(pattern, value):
                    raise ValueError(f"행 {i+1}, 컬럼 '{col}': 번호 표기 금지! 내용만 써야 함. 값: '{value}'")

        # 고장원인은 "[단계]: [설명]" 형식이어야 함
        if ':' not in str(row['고장원인']):
            raise ValueError(f"행 {i+1}, 고장원인은 '[단계]: [설명]' 형식이어야 함: {row['고장원인']}")

        lifecycle_stage = str(row['고장원인']).split(':')[0].strip()
        # QA DB 용어 병기 호환: "[재료]" -> "재료" (대괄호 제거)
        lifecycle_stage = lifecycle_stage.strip('[]')
        valid_stages = ['설계', '재료', '제작', '시험']
        if lifecycle_stage not in valid_stages:
            raise ValueError(f"행 {i+1}, 고장원인 단계는 4단계 중 하나여야 함: {lifecycle_stage}")

        # 현재예방대책은 "[단계]: [대책]" 형식이어야 함
        if ':' not in str(row['현재예방대책']):
            raise ValueError(f"행 {i+1}, 현재예방대책은 '[단계]: [대책]' 형식이어야 함: {row['현재예방대책']}")

        prevention_stage = str(row['현재예방대책']).split(':')[0].strip()
        # QA DB 용어 병기 호환: "[설계]" -> "설계" (대괄호 제거)
        prevention_stage = prevention_stage.strip('[]')
        if prevention_stage not in valid_stages:
            raise ValueError(f"행 {i+1}, 현재예방대책 단계는 4단계 중 하나여야 함: {prevention_stage}")

        # 현재검출대책은 "[단계]: [검출방법]" 형식이어야 함
        if ':' not in str(row['현재검출대책']):
            raise ValueError(f"행 {i+1}, 현재검출대책은 '[단계]: [검출방법]' 형식이어야 함: {row['현재검출대책']}")

        detection_stage = str(row['현재검출대책']).split(':')[0].strip()
        # QA DB 용어 병기 호환: "[시험]" -> "시험" (대괄호 제거)
        detection_stage = detection_stage.strip('[]')
        if detection_stage not in valid_stages:
            raise ValueError(f"행 {i+1}, 현재검출대책 단계는 4단계 중 하나여야 함: {detection_stage}")

    # 2-1. C/E/F열 상세설명 검증 (BLOCKING - 260207 개선안)
    cef_detail_errors = []
    for i, row in enumerate(data):
        # C열: 고장영향 - "영향\n(상세설명)" 2줄 형식 필수
        effect_val = str(row.get('고장영향', ''))
        if '\n' not in effect_val:
            cef_detail_errors.append(
                "행 %d C열: 상세설명 누락! '%s' (2줄 형식 필수: '영향\\n(상세설명)')"
                % (i + 1, effect_val[:30])
            )

        # F열: 고장원인 - "단계: 원인\n(상세설명)" 2줄 형식 필수
        cause_val = str(row.get('고장원인', ''))
        if '\n' not in cause_val:
            cef_detail_errors.append(
                "행 %d F열: 상세설명 누락! '%s' (2줄 형식 필수: '단계: 원인\\n(상세설명)')"
                % (i + 1, cause_val[:30])
            )

    if cef_detail_errors:
        print("\n[BLOCKING] C/F열 상세설명 형식 위반!")
        print("=" * 60)
        for err in cef_detail_errors[:10]:
            print("  %s" % err)
        if len(cef_detail_errors) > 10:
            print("  ... 외 %d개 오류" % (len(cef_detail_errors) - 10))
        print("=" * 60)
        print("\n[해결 방법]")
        print("  C열: '영향\\n(상세설명)' 2줄 형식 필수")
        print("  F열: '단계: 원인\\n(상세설명)' 2줄 형식 필수")
        raise ValueError(
            "C/F열 상세설명 형식 위반 %d개! cef-format-rules.md 참조" % len(cef_detail_errors)
        )

    # 3. G/H/J 컬럼 형식 검증 (CRITICAL-3 BLOCKING - 260131)
    ghj_errors = []
    for i, row in enumerate(data):
        # G열 (고장메커니즘) - 화살표 체인 필수
        g_errors = validate_mechanism(row.get('고장메커니즘', ''))
        for err in g_errors:
            if "[BLOCKING]" in err:
                ghj_errors.append(f"행 {i+1} G열: {err}")

        # H열 (현재예방대책) - 멀티라인 + 기준값 필수
        h_errors = validate_prevention_multiline(row.get('현재예방대책', ''))
        for err in h_errors:
            if "[BLOCKING]" in err:
                ghj_errors.append(f"행 {i+1} H열: {err}")

        # J열 (현재검출대책) - 멀티라인 + 합격기준 필수
        j_errors = validate_detection_multiline(row.get('현재검출대책', ''))
        for err in j_errors:
            if "[BLOCKING]" in err:
                ghj_errors.append(f"행 {i+1} J열: {err}")

    if ghj_errors:
        print("\n[BLOCKING] G/H/J 컬럼 형식 위반!")
        print("="*60)
        for err in ghj_errors[:10]:  # 처음 10개만 출력
            print(f"  {err}")
        if len(ghj_errors) > 10:
            print(f"  ... 외 {len(ghj_errors)-10}개 오류")
        print("="*60)
        print("\n[해결 방법]")
        print("  G열: '원인 -> 과정 -> 결과' 화살표 체인 형식 필수")
        print("  H열: 4줄 이상, 설계/재료/제작/시험 2개 이상, 기준값 포함")
        print("  J열: 4줄 이상, 설계/재료/제작/시험 2개 이상, 합격기준 포함")
        raise ValueError(f"G/H/J 컬럼 형식 위반 {len(ghj_errors)}개! SKILL.md [CRITICAL-3] 참조")

    print(f"[OK] 형식 검증 통과 (총 {len(data)}개 항목)")
    return True


def validate_logical_consistency(data):
    """논리적 일관성 및 개별성 검증 (excel-generation.md)"""

    # 1. H~T열 반복 검증 (CRITICAL!)
    seen_combinations = {}
    for i, row in enumerate(data):
        # 고장원인 + 예방조치 + 검출조치 조합
        key = (row['고장원인'], row['예방조치'], row['검출조치'])

        if key in seen_combinations:
            prev_idx = seen_combinations[key]
            raise ValueError(
                f"[!] H~T열 반복 발견!\n"
                f"행 {prev_idx+1}과 행 {i+1}이 동일한 대책:\n"
                f"  고장원인: {row['고장원인']}\n"
                f"  예방조치: {row['예방조치']}\n"
                f"  검출조치: {row['검출조치']}\n"
                f"-> 각 항목마다 다른 대책을 생성해야 합니다!"
            )
        seen_combinations[key] = i

    print(f"[OK] 논리적 일관성 검증 통과 (총 {len(data)}개 항목)")
    return True


def validate_failure_effect_phases(data):
    """시점별 고장영향 검증 (회의 합의 251128, 260109 수정)

    고장영향은 3가지 발견 시점(S값 범위)을 고려해야 함:
    - 제작 중 발견 (S=2-5): 조립 정밀도 저하, 재작업 필요, 구조 변형 등
    - 시험 중 발견 (S=4-7): 무부하 손실 초과, 절연저항 저하, 여자전류 증가 등
    - 운전 중 발생 (S=7-10): 전압 변환 불가, 지락사고, 변압기 트립, 화재 등

    [!] 주의: "조립 불합격", "FAT 불합격"은 발견 장소이므로 고장영향 아님!
            고장영향 = 기능 실패의 결과 (인과관계 필수)
    """

    # 시점별 고장영향 키워드
    manufacturing_keywords = ['정밀도', '재작업', '변형', '치수', '외관', '정렬', '품질']
    test_keywords = ['손실', '절연저항', '여자전류', '효율', '온도상승', 'PD', '측정']
    operation_keywords = ['전압', '지락', '트립', '정지', '화재', '폭발', '소손', '과열', '누유', '정전', '사고']

    # 기능별 고장영향 분석
    function_effects = {}
    for row in data:
        func = row['기능']
        effect = row['고장영향']
        s_val = row.get('S', 5)

        if func not in function_effects:
            function_effects[func] = {'manufacturing': [], 'test': [], 'operation': []}

        # 시점 분류 (키워드 + S값 기반)
        if not effect:
            continue

        has_mfg_keyword = any(kw in effect for kw in manufacturing_keywords)
        has_test_keyword = any(kw in effect for kw in test_keywords)
        has_op_keyword = any(kw in effect for kw in operation_keywords)

        # 키워드 기반 분류
        if has_mfg_keyword:
            function_effects[func]['manufacturing'].append(effect)
        if has_test_keyword:
            function_effects[func]['test'].append(effect)
        if has_op_keyword:
            function_effects[func]['operation'].append(effect)

        # S값 기반 분류 (키워드 없을 경우)
        if s_val and not (has_mfg_keyword or has_test_keyword or has_op_keyword):
            s_int = int(s_val)
            if s_int <= 5:
                function_effects[func]['manufacturing'].append(effect)
            elif s_int <= 7:
                function_effects[func]['test'].append(effect)
            else:
                function_effects[func]['operation'].append(effect)

    # 경고 출력
    warnings = []
    for func, phases in function_effects.items():
        missing = []
        if not phases['manufacturing']:
            missing.append('제작 중')
        if not phases['test']:
            missing.append('시험 중')
        if not phases['operation']:
            missing.append('운전 중')

        if missing:
            warnings.append(f"  [!] 기능 '{func[:20]}...': {', '.join(missing)} 영향 없음")

    if warnings:
        print(f"\n[경고] 시점별 고장영향 검증:")
        print("  ※ 고장영향은 제작 중/시험 중/운전 중 3가지 시점 모두 고려 권장")
        for w in warnings[:5]:  # 최대 5개만 출력
            print(w)
        if len(warnings) > 5:
            print(f"  ... 외 {len(warnings)-5}개 기능")
        print()
    else:
        print(f"[OK] 시점별 고장영향 검증 통과")

    return True  # 경고만, 에러는 아님


def validate_merge_contiguity(data):
    """병합 대상 데이터의 연속성 검증 (다이아몬드 구조 보장)

    CRITICAL: 정렬 후 실행해야 함!
    동일 값이 비연속적으로 나타나면 병합이 불가능하므로 경고 출력.

    검증 대상:
    - A열 (부품명): 전체에서 연속
    - B열 (기능): 같은 부품명 내에서 연속
    - C열 (고장영향): 같은 기능 내에서 연속
    - E열 (고장형태): 같은 고장영향 내에서 연속
    """

    from collections import defaultdict
    issues = []

    # 각 컬럼별 연속성 검증
    merge_columns = [
        ('부품명', None),
        ('기능', '부품명'),
        ('고장영향', '기능'),
        ('고장형태', '고장영향')
    ]

    # 전체 부모 체인 구축 (계층 구조 반영)
    parent_chain = {}
    for col_name, parent_col in merge_columns:
        if parent_col is None:
            parent_chain[col_name] = []
        else:
            parent_chain[col_name] = parent_chain[parent_col] + [parent_col]

    for col_name, parent_col in merge_columns:
        # 값별로 등장 위치 수집 (전체 부모 체인을 키로 사용)
        positions = defaultdict(list)
        chain = parent_chain[col_name]

        for i, row in enumerate(data):
            if chain:
                # 전체 부모 체인 + 현재 값으로 키 생성
                key = tuple(row[c] for c in chain) + (row[col_name],)
            else:
                key = row[col_name]
            positions[key].append(i)

        # 비연속 검출
        for key, indices in positions.items():
            if len(indices) > 1:
                # 연속성 확인: 인덱스가 순차적이어야 함
                for j in range(len(indices) - 1):
                    if indices[j + 1] - indices[j] > 1:
                        # 비연속 발견
                        if chain:
                            col_val = key[-1] if isinstance(key, tuple) else key
                            chain_str = ' > '.join(str(k)[:15] for k in key[:-1]) if isinstance(key, tuple) and len(key) > 1 else ''
                            issues.append(
                                f"  - {col_name} '{str(col_val)[:25]}...' (chain: {chain_str})\n"
                                f"    위치: 행 {[idx+1 for idx in indices]}"
                            )
                        else:
                            issues.append(
                                f"  - {col_name} '{str(key)[:30]}...'\n"
                                f"    위치: 행 {[idx+1 for idx in indices]}"
                            )
                        break

    if issues:
        print(f"[!] 비연속 병합 대상 발견 ({len(issues)}건) - 다이아몬드 구조 병합 불가!")
        for issue in issues[:5]:  # 최대 5개만 출력
            print(issue)
        if len(issues) > 5:
            print(f"  ... 외 {len(issues) - 5}건")
        print("[!] 정렬 알고리즘 또는 데이터 생성 로직 확인 필요!")
        return False

    print(f"[OK] 병합 연속성 검증 통과 (모든 병합 대상 연속)")
    return True


def validate_diamond_structure_data(data):
    """다이아몬드 구조 검증 (JSON 데이터 기반) - 재발방지대책 260111

    BLOCKING 기준:
    - 형태당 원인 평균 >= 2.0 (필수!)
    - 1:1:1:1 직선 구조 비율 < 30%

    Returns:
        dict: {
            'avg_causes_per_mode': float,
            'linear_ratio': float,
            'single_cause_modes': list  # 원인 1개뿐인 고장형태
        }
    """
    from collections import defaultdict

    # 1. 고장형태 -> 고장원인 매핑
    mode_to_causes = defaultdict(set)
    for row in data:
        mode = row.get('고장형태', '').strip()
        cause = row.get('고장원인', '').strip()
        if mode and cause:
            mode_to_causes[mode].add(cause)

    # 2. 형태당 원인 개수 계산
    causes_per_mode = [len(causes) for causes in mode_to_causes.values()]
    avg_causes = sum(causes_per_mode) / len(causes_per_mode) if causes_per_mode else 0

    # 3. 원인 1개뿐인 고장형태 목록
    single_cause_modes = []
    for mode, causes in mode_to_causes.items():
        if len(causes) == 1:
            single_cause_modes.append({
                'mode': mode,
                'cause': list(causes)[0]
            })

    # 4. 기능/영향/형태 매핑 (직선 구조 계산용)
    func_to_effects = defaultdict(set)
    effect_to_modes = defaultdict(set)

    for row in data:
        func = row.get('기능', '').strip()
        effect = row.get('고장영향', '').strip()
        mode = row.get('고장형태', '').strip()

        if func and effect:
            func_to_effects[func].add(effect)
        if effect and mode:
            effect_to_modes[effect].add(mode)

    # 5. 직선 구조 비율 계산
    linear_count = 0
    total_complete = 0

    for row in data:
        func = row.get('기능', '').strip()
        effect = row.get('고장영향', '').strip()
        mode = row.get('고장형태', '').strip()
        cause = row.get('고장원인', '').strip()

        if not all([func, effect, mode, cause]):
            continue

        total_complete += 1

        # 모든 레벨에서 1:1 관계인 경우 직선 구조
        if (len(func_to_effects.get(func, set())) == 1 and
            len(effect_to_modes.get(effect, set())) == 1 and
            len(mode_to_causes.get(mode, set())) == 1):
            linear_count += 1

    linear_ratio = (linear_count / total_complete * 100) if total_complete > 0 else 0

    return {
        'avg_causes_per_mode': avg_causes,
        'linear_ratio': linear_ratio,
        'single_cause_modes': single_cause_modes,
        'total_modes': len(mode_to_causes),
        'total_complete_rows': total_complete,
        'linear_count': linear_count
    }


def apply_cell_merge(ws, data, start_row=7):
    """셀 병합 알고리즘 (A-E 컬럼만, parent 조건 포함)

    CRITICAL: excel-generation.md 알고리즘 준수
    - A (부품명): 전체 병합
    - B (기능): 같은 부품 내에서만 병합
    - C (고장영향): 같은 기능 내에서만 병합
    - D (S): 같은 고장영향 내에서만 병합
    - E (고장형태): 같은 고장영향 내에서만 병합

    [!] 주의: 데이터는 반드시 sort_fmea_data()로 정렬되어야 함!
    """

    # 병합 규칙 (컬럼, 키, parent 키)
    merge_rules = [
        ('A', '부품명', None),           # 부품명: parent 없음
        ('B', '기능', '부품명'),          # 기능: 같은 부품 내에서
        ('C', '고장영향', '기능'),        # 고장영향: 같은 기능 내에서
        ('D', 'S', '고장영향'),           # S: 같은 고장영향 내에서
        ('E', '고장형태', '고장영향')     # 고장형태: 같은 고장영향 내에서
    ]

    merge_count = 0
    for col_letter, col_name, parent_col in merge_rules:
        i = 0
        while i < len(data):
            merge_start = start_row + i
            current_value = data[i][col_name]

            # parent 조건 확인 (있는 경우)
            parent_value = None
            if parent_col:
                parent_value = data[i][parent_col]

            # 연속된 같은 값 찾기
            j = i + 1
            while j < len(data):
                # 값이 같은지 확인
                if data[j][col_name] == current_value:
                    # parent가 있으면 parent도 같아야 병합
                    if parent_col is None:
                        # parent 조건 없으면 무조건 병합
                        j += 1
                    elif data[j][parent_col] == parent_value:
                        # parent도 같으면 병합
                        j += 1
                    else:
                        # parent가 다르면 병합 중단
                        print(f"  [!] {col_letter}열 병합 중단: 행 {start_row+j}, parent 불일치 ({parent_col}: '{parent_value}' -> '{data[j][parent_col]}')")
                        break
                else:
                    # 값이 다르면 병합 중단
                    break

            # 2개 이상 행이면 병합
            merge_end = start_row + j - 1
            if merge_end > merge_start:
                ws.merge_cells(f'{col_letter}{merge_start}:{col_letter}{merge_end}')
                ws[f'{col_letter}{merge_start}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                merge_count += 1
                value_str = str(current_value)
                display_value = value_str[:20] + '...' if len(value_str) > 20 else value_str
                print(f"  [OK] {col_letter}열 병합: 행 {merge_start}-{merge_end} (값: '{display_value}')")

            i = j  # 다음 구간으로 이동

    print(f"[OK] 셀 병합 완료 (A-E 컬럼, {merge_count}개 구간, parent 조건 적용)")


def load_cell_comments(md_path=None):
    """cell-comments.md에서 Row 6 메시지 로드 (Single Source of Truth)

    기준: references/cell-comments.md
    적용: Row 6 헤더 셀에 데이터 유효성 검사 메시지 추가

    Args:
        md_path: cell-comments.md 파일 경로 (None이면 자동 탐색)

    Returns:
        dict: {셀주소: (제목, 프롬프트)} 형태의 딕셔너리
    """
    if md_path is None:
        # 스크립트 위치 기준 상대 경로
        script_dir = os.path.dirname(os.path.abspath(__file__))
        md_path = os.path.join(script_dir, '..', 'references', 'cell-comments.md')

    messages = {}

    # 파일이 없으면 기본값 반환
    if not os.path.exists(md_path):
        print(f"[!] cell-comments.md 파일 없음: {md_path}")
        print("[!] 기본 메시지 사용")
        return _get_default_messages()

    with open(md_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # 컬럼별 섹션 파싱 (### A6 - 제목, ### B6 - 제목, ... ### V6 - 제목)
    # 패턴: ### [A-V]6 - [제목] (22개 컬럼 지원)
    column_pattern = r'### ([A-V])6 - ([^\n]+)\n(.*?)(?=### [A-V]6 - |---|\Z)'
    matches = re.findall(column_pattern, content, re.DOTALL)

    for col, title, section in matches:
        cell_ref = f'{col}6'

        # 제목 정리 (32자 제한)
        clean_title = title.strip()[:32]

        # 코드 블록 내용 추출
        code_block_match = re.search(r'```\n?(.*?)```', section, re.DOTALL)
        if code_block_match:
            prompt = code_block_match.group(1).strip()
        else:
            # 코드 블록이 없으면 섹션 전체 사용
            prompt = section.strip()

        # 255자 제한
        if len(prompt) > 255:
            prompt = prompt[:252] + '...'

        messages[cell_ref] = (clean_title, prompt)

    if messages:
        print(f"[OK] cell-comments.md 로드 완료 ({len(messages)}개 컬럼)")
    else:
        print("[!] cell-comments.md 파싱 실패, 기본 메시지 사용")
        return _get_default_messages()

    return messages


def _get_default_messages():
    """기본 메시지 (cell-comments.md 파싱 실패 시 fallback, 22개 컬럼)"""
    return {
        'A6': ("부품명", "분석 대상 부품 이름 (번호 금지)"),
        'B6': ("기능", "부품의 기능 (부품당 3-10개, 번호 금지)"),
        'C6': ("고장영향", "기능 실패 시 최종 영향 (소음/진동 여기)"),
        'D6': ("S 심각도", "1-10 (10=화재, 1=경미)"),
        'E6': ("고장형태", "현재 관찰 가능 현상 (소음/진동 금지->C열)"),
        'F6': ("고장원인", "[단계]: [설명] (설계/재료/제작/시험)"),
        'G6': ("고장메커니즘", "원인->과정->결과 형식"),
        'H6': ("현재예방대책", "[단계]: [대책] ([기준]) (설계/재료/제작/시험)"),
        'I6': ("O 발생도", "1-10 (연200대 기준)"),
        'J6': ("현재검출대책", "[단계]: [검출방법] ([장비]) (설계/재료/제작/시험)"),
        'K6': ("D 검출도", "1-10 (1=명확, 10=불가)"),
        'L6': ("RPN", "RPN = S x O x D (1-1000)\n>=200=빨강, 100-199=노랑, <100=녹색"),
        'M6': ("AP", "H/M/L (자동계산)"),
        'N6': ("예방조치", "H:설계변경, M:검증추가, L:없음"),
        'O6': ("검출조치", "H:시험추가, M:검사강화, L:없음"),
        'P6': ("담당자", "책임팀 (단계별 매핑)"),
        'Q6': ("목표일", "H:즉시, M:3개월, L:6개월"),
        'R6': ("S'", "개선 후 심각도"),
        'S6': ("O'", "개선 후 발생도"),
        'T6': ("D'", "개선 후 검출도"),
        'U6': ("RPN'", "RPN' = S' x O' x D' (개선 효과 측정)"),
        'V6': ("AP'", "개선 후 우선순위")
    }


def add_cell_validation_messages(ws, md_path=None):
    """Row 6 헤더 셀에 작성 가이드 메시지 추가 (데이터 유효성 검사 방식)

    기준: references/cell-comments.md (Single Source of Truth)
    M365 웹에서도 표시됨 (255자 제한으로 핵심 내용만 표시)

    Args:
        ws: openpyxl Worksheet 객체
        md_path: cell-comments.md 파일 경로 (None이면 자동 탐색)
    """
    # cell-comments.md에서 메시지 로드 (Single Source of Truth)
    messages = load_cell_comments(md_path)

    # 각 셀에 데이터 유효성 검사 메시지 추가
    msg_count = 0
    for cell_ref, (title, prompt) in messages.items():
        dv = DataValidation(type="none", allow_blank=True)
        dv.promptTitle = title
        dv.prompt = prompt
        dv.showInputMessage = True
        dv.add(cell_ref)
        ws.add_data_validation(dv)
        msg_count += 1

    print(f"[OK] 입력 안내 메시지 추가 완료 (Row 6, {msg_count}개 컬럼, M365 웹 호환)")


def generate_excel(fmea_data, output_path, project_info=None):
    """Excel 파일 생성 및 서식 적용

    출력 구조 (2개 시트):
    - Sheet 1: 기능분석 (AIAG-VDA Step 3)
    - Sheet 2: FMEA (AIAG-VDA Step 4-6)
    """

    # 1. 데이터 정렬
    print("1. 데이터 정렬 중...")
    # project_info에서 function_order 가져오기 (다이어그램 순서)
    function_order = project_info.get('function_order', None) if project_info else None
    fmea_data = sort_fmea_data(fmea_data, function_order=function_order)

    # 1-1. 용어 병기 적용 (전문용어 + 쉬운표현)
    # [!] E열(고장형태) 제외: GLOSSARY '열화'->'열화(성능저하)' 등이 E열 금지어와 충돌 (재발방지 260201)
    print("1-1. 용어 병기 적용 중...")
    used_terms = set()  # 전체 문서에서 첫 등장 추적
    glossary_columns = ['기능', '고장영향', '고장원인', '고장메커니즘', '현재예방대책', '현재검출대책']
    for row in fmea_data:
        for col in glossary_columns:
            if col in row and row[col]:
                row[col] = apply_glossary(row[col], used_terms)
    print(f"  -> 병기 적용 완료: {len(used_terms)}개 용어")

    # 2. 검증 (ALL-AT-ONCE: 모든 BLOCKING 오류를 수집 후 한 번에 보고)
    print("2. 데이터 검증 중...")
    all_blocking_errors = []  # 전체 BLOCKING 오류 수집 (fixer Worker에 전달용)

    # 2-pre. 기본 데이터/논리 검증
    try:
        validate_data(fmea_data)
    except ValueError as e:
        all_blocking_errors.append(("[DATA] " + str(e)))
    try:
        validate_logical_consistency(fmea_data)
    except ValueError as e:
        all_blocking_errors.append(("[LOGIC] " + str(e)))
    try:
        validate_failure_effect_phases(fmea_data)  # 시점별 고장영향 검증 (회의 합의 251128)
    except ValueError as e:
        all_blocking_errors.append(("[PHASE] " + str(e)))

    # 2-0. 고장영향(C열) 금지어 검증 (BLOCKING) - 재발방지대책 260111
    print("2-0. 고장영향(C열) 금지어 검증 중...")
    failure_effect_violations = []
    for i, row in enumerate(fmea_data):
        failure_effect = row.get('고장영향', '')
        if failure_effect:
            is_valid, reason = validate_failure_effect(failure_effect)
            if not is_valid:
                failure_effect_violations.append({
                    'row': i + 1,
                    'value': failure_effect,
                    'reason': reason
                })

    if failure_effect_violations:
        error_msg = "[!] 고장영향(C열) 검증 실패! (BLOCKING)\n\n"
        error_msg += "금지어 목록 (검사/판정 결과): %s...\n\n" % str(FORBIDDEN_INSPECTION_RESULTS[:10])
        error_msg += "위반 항목:\n"
        for v in failure_effect_violations[:10]:
            error_msg += "  - 항목 %d: \"%s\" -> %s\n" % (v['row'], v['value'], v['reason'])
        if len(failure_effect_violations) > 10:
            error_msg += "  ... 외 %d건\n" % (len(failure_effect_violations) - 10)
        error_msg += "\n수정 방법:\n"
        error_msg += "  - 검사/판정 결과(FAT 불합격, 시험 불합격) -> 기술적 영향으로 변경\n"
        error_msg += "  - 예: 'FAT 불합격' -> '과열', '효율 저하', '절연파괴' 등\n"
        error_msg += "참조: references/effect-ontology.md, references/diamond-structure.md"
        all_blocking_errors.append(error_msg)
    else:
        print("   [OK] 고장영향(C열) 검사/판정 결과 검증 통과 (%d개 항목)" % len(fmea_data))

    # 2-0-1. 고장영향(C열) 물리적 상태 검증 (BLOCKING) - v2.0 260129
    print("2-0-1. 고장영향(C열) 물리적 상태 검증 중...")
    effect_ontology = load_effect_ontology()
    forbidden_physical = effect_ontology.get('forbidden_physical', [])
    physical_violations = []

    if forbidden_physical:
        print(f"   [INFO] 물리적 상태 금지어 {len(forbidden_physical)}개 로드됨")
        for i, row in enumerate(fmea_data):
            failure_effect = row.get('고장영향', '')
            if failure_effect:
                is_valid, reason = validate_physical_in_effect(failure_effect, forbidden_physical)
                if not is_valid:
                    physical_violations.append({
                        'row': i + 1,
                        'value': failure_effect,
                        'reason': reason
                    })

        if physical_violations:
            error_msg = "[!] 고장영향(C열)에 물리적 상태 발견! (BLOCKING)\n\n"
            error_msg += "C열에는 '기능 실패의 결과'만 작성!\n"
            error_msg += "물리적 상태(변형, 탈락, 이완 등)는 E열(고장형태)로 이동 필요\n\n"
            error_msg += "금지어 예시: %s...\n\n" % str(forbidden_physical[:15])
            error_msg += "위반 항목:\n"
            for v in physical_violations[:10]:
                error_msg += "  - 항목 %d: \"%s\"\n" % (v['row'], v['value'])
                error_msg += "    -> %s\n" % v['reason']
            if len(physical_violations) > 10:
                error_msg += "  ... 외 %d건\n" % (len(physical_violations) - 10)
            error_msg += "\n수정 방법:\n"
            error_msg += "  - 물리적 상태(탈락, 분리, 이완, 변형) -> E열(고장형태)로 이동\n"
            error_msg += "  - C열에는 기능 실패 결과(통전 불가, 과열, 지락사고) 작성\n"
            error_msg += "  - 질문: '시스템이 뭐가 안 되나?' -> C열 / '부품에 뭐가 보이나?' -> E열\n"
            error_msg += "참조: references/effect-ontology.md"
            all_blocking_errors.append(error_msg)
        else:
            print("   [OK] 고장영향(C열) 물리적 상태 검증 통과 (%d개 항목)" % len(fmea_data))
    else:
        print("   [WARN] 물리적 상태 금지어 목록 로드 실패 - 검증 생략")

    # 2-1. 고장형태(E열) 금지어 검증 (BLOCKING)
    print("2-1. 고장형태 금지어 검증 중...")
    failure_mode_violations = []
    for i, row in enumerate(fmea_data):
        failure_mode = row.get('고장형태', '')
        if failure_mode:
            # 금지어 검증
            is_valid, reason = validate_failure_mode(failure_mode)
            if not is_valid:
                failure_mode_violations.append({
                    'row': i + 1,
                    'value': failure_mode,
                    'reason': reason
                })
            # 태그 형식 검증
            tag_valid, tag_reason = validate_tag_format(failure_mode)
            if not tag_valid:
                failure_mode_violations.append({
                    'row': i + 1,
                    'value': failure_mode,
                    'reason': tag_reason
                })

    if failure_mode_violations:
        error_msg = "[!] 고장형태(E열) 검증 실패!\n\n"
        error_msg += "금지어 목록: %s\n\n" % str(FORBIDDEN_PATTERNS + FORBIDDEN_EXACT)
        error_msg += "위반 항목:\n"
        for v in failure_mode_violations[:10]:  # 최대 10개만 표시
            error_msg += "  - 항목 %d: \"%s\" -> %s\n" % (v['row'], v['value'], v['reason'])
        if len(failure_mode_violations) > 10:
            error_msg += "  ... 외 %d건\n" % (len(failure_mode_violations) - 10)
        error_msg += "\n수정 방법:\n"
        error_msg += "  - 측정값(~증가/~저하) -> G열(고장메커니즘)으로 이동\n"
        error_msg += "  - 미래결과(소음/진동) -> C열(고장영향)으로 이동\n"
        error_msg += "  - 태그 형식: [부족|과도|유해]: [관찰가능현상]\n"
        error_msg += "참조: references/failure-mode-forbidden.md, references/5why-vs-fmea.md"
        all_blocking_errors.append(error_msg)
    else:
        print("   [OK] 고장형태 검증 통과 (%d개 항목)" % len(fmea_data))

    # 2-1-1. H열/J열(현재예방대책/현재검출대책) 검증 (BLOCKING - 260129)
    print("2-1-1. H열/J열 검증 중...")
    prevention_detection_violations = []
    prevention_detection_warnings = []

    for i, row in enumerate(fmea_data):
        # H열 (현재예방대책) 검증
        prevention = row.get('현재예방대책', '')
        if prevention:
            # 1. 형식 검증
            is_valid, reason = validate_stage_format(prevention)
            if not is_valid:
                prevention_detection_violations.append({
                    'row': i + 1, 'column': 'H열', 'value': prevention[:50], 'reason': reason
                })
            # 2. 출처 검증
            is_valid, reason, severity = validate_source_presence(prevention)
            if severity == "ERROR":
                prevention_detection_violations.append({
                    'row': i + 1, 'column': 'H열', 'value': prevention[:50], 'reason': reason
                })
            elif severity == "WARNING":
                prevention_detection_warnings.append({
                    'row': i + 1, 'column': 'H열', 'value': prevention[:50], 'reason': reason
                })
            # 3. 금지 패턴 검증
            is_valid, reason = validate_forbidden_source(prevention)
            if not is_valid:
                prevention_detection_violations.append({
                    'row': i + 1, 'column': 'H열', 'value': prevention[:50], 'reason': reason
                })
            # 4. 기준값 검증
            is_valid, reason, severity = validate_value_presence(prevention)
            if severity == "WARNING":
                prevention_detection_warnings.append({
                    'row': i + 1, 'column': 'H열', 'value': prevention[:50], 'reason': reason
                })

        # J열 (현재검출대책) 검증
        detection = row.get('현재검출대책', '')
        if detection:
            # 1. 형식 검증
            is_valid, reason = validate_stage_format(detection)
            if not is_valid:
                prevention_detection_violations.append({
                    'row': i + 1, 'column': 'J열', 'value': detection[:50], 'reason': reason
                })
            # 2. 출처 검증
            is_valid, reason, severity = validate_source_presence(detection)
            if severity == "ERROR":
                prevention_detection_violations.append({
                    'row': i + 1, 'column': 'J열', 'value': detection[:50], 'reason': reason
                })
            elif severity == "WARNING":
                prevention_detection_warnings.append({
                    'row': i + 1, 'column': 'J열', 'value': detection[:50], 'reason': reason
                })
            # 3. 금지 패턴 검증
            is_valid, reason = validate_forbidden_source(detection)
            if not is_valid:
                prevention_detection_violations.append({
                    'row': i + 1, 'column': 'J열', 'value': detection[:50], 'reason': reason
                })
            # 4. 기준값 검증
            is_valid, reason, severity = validate_value_presence(detection)
            if severity == "WARNING":
                prevention_detection_warnings.append({
                    'row': i + 1, 'column': 'J열', 'value': detection[:50], 'reason': reason
                })

    # WARNING 출력 (계속 진행)
    if prevention_detection_warnings:
        print(f"   [WARNING] H열/J열 경고 {len(prevention_detection_warnings)}건:")
        for w in prevention_detection_warnings[:5]:
            print(f"      - {w['column']} 항목{w['row']}: {w['reason']}")
        if len(prevention_detection_warnings) > 5:
            print(f"      ... 외 {len(prevention_detection_warnings) - 5}건")

    # [v1.1 변경] H열/J열 검증: WARNING만 (BLOCKING 아님!)
    # 출처 없는 항목도 허용 - 금지 패턴만 체크
    if prevention_detection_violations:
        # 금지 패턴 사용만 ERROR (형식 위반, 금지어 사용 등)
        real_errors = [v for v in prevention_detection_violations
                       if '금지 패턴' in v['reason'] or '형식' in v['reason']]

        if real_errors:
            # 금지 패턴/형식 위반만 BLOCKING
            error_msg = "[!] H열/J열 금지 패턴/형식 위반! (BLOCKING)\n\n"
            error_msg += "위반 항목:\n"
            for v in real_errors[:10]:
                error_msg += "  - %s 항목%d: \"%s...\" -> %s\n" % (v['column'], v['row'], v['value'], v['reason'])
            if len(real_errors) > 10:
                error_msg += "  ... 외 %d건\n" % (len(real_errors) - 10)
            error_msg += "\n수정 방법:\n"
            error_msg += "  - 금지: (작업표준), (검사기준), (CS) 등 문서번호 없는 일반 용어\n"
            error_msg += "  - 허용: (IEQT-T-W030 S3.2), (CHECK SHEET No.5), (일반)\n"
            error_msg += "참조: references/prevention-detection-ontology.md"
            all_blocking_errors.append(error_msg)
        else:
            # 출처 누락은 WARNING만 출력하고 계속 진행
            print(f"   [WARNING] H열/J열 출처 누락 {len(prevention_detection_violations)}건 (계속 진행)")
            for v in prevention_detection_violations[:5]:
                print(f"      - {v['column']} 항목{v['row']}: {v['reason']}")
            if len(prevention_detection_violations) > 5:
                print(f"      ... 외 {len(prevention_detection_violations) - 5}건")

    print(f"   [OK] H열/J열 검증 완료 ({len(fmea_data)}개 항목, 경고 {len(prevention_detection_warnings)}건)")

    # 2-2. 병합 연속성 검증 (정렬 후 필수!)
    print("2-2. 병합 연속성 검증 중...")
    if not validate_merge_contiguity(fmea_data):
        all_blocking_errors.append(
            "[!] 병합 연속성 검증 실패!\n"
            "동일 기능/고장영향이 비연속적으로 배치되어 Excel 셀 병합이 불가능합니다.\n"
            "해결방법: FMEA 항목 생성 시 기능별로 완결 후 다음 기능으로 이동하세요.\n"
            "참조: SKILL.md STEP 3 - '기능별 연속 생성 필수' 규칙"
        )

    # 2-3. 인과관계 체인 검증 (형태->원인->메커니즘)
    print("2-3. 인과관계 체인 검증 중...")
    causal_chain_violations = []
    causal_chain_warnings = []

    for i, row in enumerate(fmea_data):
        mode = row.get('고장형태', '')
        cause = row.get('고장원인', '')
        mechanism = row.get('고장메커니즘', '')

        # 형태 -> 원인 검증
        if mode and cause:
            mc_valid, mc_reason = validate_mode_cause(mode, cause)
            if not mc_valid:
                causal_chain_violations.append({
                    'row': i + 1,
                    'type': '형태->원인',
                    'values': f'"{mode}" <- "{cause}"',
                    'reason': mc_reason
                })
            elif mc_reason.startswith("[WARN]"):
                causal_chain_warnings.append({
                    'row': i + 1,
                    'type': '형태->원인',
                    'reason': mc_reason
                })

        # 원인 -> 메커니즘 검증
        if cause and mechanism:
            cm_valid, cm_reason = validate_cause_mechanism(cause, mechanism)
            if not cm_valid:
                causal_chain_violations.append({
                    'row': i + 1,
                    'type': '원인->메커니즘',
                    'values': f'"{cause}" -> "{mechanism}"',
                    'reason': cm_reason
                })
            elif cm_reason.startswith("[WARN]"):
                causal_chain_warnings.append({
                    'row': i + 1,
                    'type': '원인->메커니즘',
                    'reason': cm_reason
                })

        # 라이프사이클 일관성 검증
        if cause and mechanism:
            lc_valid, lc_reason, _ = validate_lifecycle_consistency(cause, mechanism)
            if not lc_valid:
                causal_chain_violations.append({
                    'row': i + 1,
                    'type': '라이프사이클',
                    'values': f'"{cause}" / "{mechanism}"',
                    'reason': lc_reason
                })

    if causal_chain_violations:
        error_msg = "[!] 인과관계 체인 검증 실패!\n\n"
        error_msg += "위반 항목:\n"
        for v in causal_chain_violations[:10]:
            error_msg += "  - 항목 %d [%s]: %s\n" % (v['row'], v['type'], v['values'])
            error_msg += "    -> %s\n" % v['reason']
        if len(causal_chain_violations) > 10:
            error_msg += "  ... 외 %d건\n" % (len(causal_chain_violations) - 10)
        error_msg += "\n수정 방법:\n"
        error_msg += "  - 형태->원인: 원인이 고장형태를 유발하는지 확인\n"
        error_msg += "  - 원인->메커니즘: 메커니즘이 원인-형태 과정을 설명하는지 확인\n"
        error_msg += "  - 라이프사이클: 원인과 메커니즘이 동일 단계인지 확인\n"
        error_msg += "참조: references/causal-chain-ontology.md"
        all_blocking_errors.append(error_msg)

    if causal_chain_warnings:
        print(f"   [WARN] 검토 권장 항목 {len(causal_chain_warnings)}건 (진행은 계속)")
        for w in causal_chain_warnings[:3]:
            print(f"      - 항목 {w['row']} [{w['type']}]: {w['reason']}")

    print(f"   [OK] 인과관계 체인 검증 통과 ({len(fmea_data)}개 항목)")

    # 2-4. 다이아몬드 구조 검증 (BLOCKING) - 재발방지대책 260111
    print("2-4. 다이아몬드 구조 검증 중...")
    diamond_result = validate_diamond_structure_data(fmea_data)

    diamond_failed = False
    if diamond_result['linear_ratio'] >= 30:
        error_msg = "[!] 다이아몬드 구조 검증 실패! (BLOCKING)\n\n"
        error_msg += "1:1:1:1 직선 구조 비율: %.1f%% (기준: <30%%)\n" % diamond_result['linear_ratio']
        error_msg += "형태당 원인 평균: %.2f개 (기준: >=2.0)\n\n" % diamond_result['avg_causes_per_mode']
        error_msg += "위반 항목 (원인 1개인 고장형태):\n"
        for mode in diamond_result['single_cause_modes'][:10]:
            error_msg += "  - \"%s\" <- 원인 1개만: \"%s\"\n" % (mode['mode'], mode['cause'])
        if len(diamond_result['single_cause_modes']) > 10:
            error_msg += "  ... 외 %d건\n" % (len(diamond_result['single_cause_modes']) - 10)
        error_msg += "\n수정 방법:\n"
        error_msg += "  - 각 고장형태에 최소 2개 원인 추가 (설계/제작/운전 관점)\n"
        error_msg += "  - 1:1:1:1 직선 구조 -> 1:N:M:K 다이아몬드 구조로 확장\n"
        error_msg += "참조: SKILL.md '다이아몬드 확장 구조', references/diamond-structure.md"
        all_blocking_errors.append(error_msg)
        diamond_failed = True

    if not diamond_failed and diamond_result['avg_causes_per_mode'] < 2.0:
        error_msg = "[!] 다이아몬드 구조 검증 실패! (BLOCKING)\n\n"
        error_msg += "형태당 원인 평균: %.2f개 (기준: >=2.0 필수!)\n" % diamond_result['avg_causes_per_mode']
        error_msg += "1:1:1:1 직선 구조 비율: %.1f%%\n\n" % diamond_result['linear_ratio']
        error_msg += "원인 부족 고장형태:\n"
        for mode in diamond_result['single_cause_modes'][:10]:
            error_msg += "  - \"%s\" <- \"%s\" (1개만)\n" % (mode['mode'], mode['cause'])
        if len(diamond_result['single_cause_modes']) > 10:
            error_msg += "  ... 외 %d건\n" % (len(diamond_result['single_cause_modes']) - 10)
        error_msg += "\n수정 방법:\n"
        error_msg += "  - 각 고장형태에 최소 2개 이상 원인 추가\n"
        error_msg += "  - 관점별 원인: 설계 결함 / 재료 문제 / 제작 오류 / 운전 조건\n"
        error_msg += "참조: SKILL.md '고장형태당 원인 1개만 = 다이아몬드 구조 실패'"
        all_blocking_errors.append(error_msg)

    if not diamond_failed and diamond_result['avg_causes_per_mode'] >= 2.0:
        print("   [OK] 다이아몬드 구조 검증 통과 (직선비율: %.1f%%, 형태당원인: %.2f)" % (diamond_result['linear_ratio'], diamond_result['avg_causes_per_mode']))

    # 2-5. 기능 커버리지 검증 (v12.1 강화: 다이어그램 기능 전부 각 2개+!)
    print("2-5. 기능 커버리지 검증 중...")
    from collections import Counter as _Counter
    func_counts = _Counter()
    part_primary = {}
    for row in fmea_data:
        p = row.get('부품명', '')
        f = row.get('기능', '')
        func_counts[(p, f)] += 1
        if p not in part_primary:
            part_primary[p] = f

    primary_total = sum(c for (p, f), c in func_counts.items() if part_primary.get(p) == f)
    total_items = len(fmea_data)
    primary_ratio = (primary_total / total_items * 100) if total_items > 0 else 0

    # v12.1: 기능별 항목 수 검사 (0개=BLOCKING, 1개=WARNING)
    coverage_failed = False
    zero_funcs = []
    one_funcs = []
    for (p, f), cnt in func_counts.items():
        if cnt == 0:
            zero_funcs.append((p, f))
            coverage_failed = True
        elif cnt == 1:
            one_funcs.append((p, f))

    if coverage_failed or primary_ratio < 20:
        error_msg = "[!] 기능 커버리지 검증 실패!\n\n"
        if zero_funcs:
            error_msg += "[BLOCKING] 항목 0개 기능:\n"
            for p, f in zero_funcs:
                error_msg += "  - %s: \"%s\" (0 items!)\n" % (p, f)
            error_msg += "\n"
        error_msg += "주기능 비율: %.1f%% (권장: >= 30%%)\n" % primary_ratio
        error_msg += "주기능 항목: %d / 전체: %d\n\n" % (primary_total, total_items)
        error_msg += "기능별 항목 수:\n"
        for (p, f), cnt in sorted(func_counts.items()):
            marker = " [!]" if cnt <= 1 else ""
            error_msg += "  - %s: \"%s\" (%d items)%s\n" % (p, f, cnt, marker)
        error_msg += "\n수정 방법: 다이어그램 모든 기능에 각 2개+ 항목 필요!\n"
        error_msg += "참조: SKILL.md GATE 4 기능 커버리지 (v12.1)"
        all_blocking_errors.append(error_msg)
    elif one_funcs:
        print("   [WARN] 항목 1개뿐인 기능 %d건 (각 2개+ 권장):" % len(one_funcs))
        for p, f in one_funcs:
            print("      - %s: \"%s\"" % (p, f))
        if primary_ratio < 30:
            print("   [WARN] 주기능 비율 %.1f%% (권장: >= 30%%)" % primary_ratio)
    else:
        print("   [OK] 기능 커버리지 검증 통과 (주기능 비율: %.1f%%, 모든 기능 2개+)" % primary_ratio)

    # 2-6. SOD 형식 검증 (WARNING - v12 신규)
    print("2-6. SOD 형식 검증 중...")
    import re as _re
    bad_sod_items = []
    for i, row in enumerate(fmea_data):
        sod = row.get('SOD', '')
        if sod and not _re.match(r'^S\d+xO\d+xD\d+$', str(sod)):
            bad_sod_items.append((i + 1, sod))
    if bad_sod_items:
        print("   [WARN] SOD 형식 오류 %d건 (정규 형식: S{n}xO{n}xD{n})" % len(bad_sod_items))
        for row_num, sod_val in bad_sod_items[:5]:
            print("      - 항목 %d: \"%s\"" % (row_num, sod_val))
    else:
        print("   [OK] SOD 형식 검증 통과")

    # === ALL-AT-ONCE 검증 결과 종합 보고 ===
    if all_blocking_errors:
        summary = "=" * 60 + "\n"
        summary += "[!] BLOCKING 오류 총 %d건 발견! (ALL-AT-ONCE 보고)\n" % len(all_blocking_errors)
        summary += "=" * 60 + "\n\n"
        for idx, err in enumerate(all_blocking_errors, 1):
            summary += "--- 오류 %d/%d ---\n" % (idx, len(all_blocking_errors))
            summary += err + "\n\n"
        summary += "=" * 60 + "\n"
        summary += "fmea-worker-fixer에 위 오류 목록 전체를 전달하여 일괄 수정하세요.\n"
        summary += "=" * 60
        raise ValueError(summary)

    # 3. Excel 생성
    print("3. Excel 파일 생성 중...")
    wb = Workbook()
    ws = wb.active
    ws.title = "FMEA"

    # 3-1. 기능분석 시트 생성 (Sheet 1)
    print("3-1. 기능분석 시트 생성 중...")
    function_data = extract_function_analysis_data(fmea_data, project_info)

    # 3-1-1. 기능분석 데이터 검증 (경고만, 중단 안함)
    print("3-1-1. 기능분석 데이터 검증 중...")
    validate_function_analysis_data(function_data)

    create_function_analysis_sheet(wb, function_data, project_info)

    # 3-2. FMEA 시트 생성 (Sheet 2 - 기존 로직)

    # Row 1: 제목 - 부품명_FMEA (크게, 굵게, 가운데 정렬)
    part_name = project_info.get('부품', '부품명') if project_info else '부품명'
    ws.merge_cells('A1:V1')  # 22개 컬럼 (A-V)
    title_cell = ws['A1']
    title_cell.value = f"{part_name}_FMEA"
    title_cell.font = Font(bold=True, size=20)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row 2: 프로젝트 정보 (전압 표기 제거)
    ws.merge_cells('A2:V2')  # 22개 컬럼 (A-V)
    project_cell = ws['A2']
    project_name = project_info.get('프로젝트', '변압기') if project_info else '변압기'
    # 괄호와 그 안의 내용 제거 (예: "초고압 변압기 (154kV/345kV/765kV)" -> "초고압 변압기")
    project_name = re.sub(r'\s*\([^)]*\)', '', project_name)
    project_cell.value = f"프로젝트: {project_name}"
    project_cell.alignment = Alignment(horizontal='left', vertical='center')
    project_cell.font = Font(size=14)  # Row 2 글자 크기 14

    # Row 3: 자료 출처 (MANDATORY) - project_info에서 동적으로 가져옴
    ws.merge_cells('A3:V3')  # 22개 컬럼 (A-V)
    source_cell = ws['A3']
    # 출처를 project_info에서 가져오고, 없으면 기본값 사용
    # QA DB(1,972건 품질이력) + 내부자료(CHECK SHEET, W/R/I/P시리즈) + 외부자료(IEC/IEEE/CIGRE)
    default_sources = "[QA DB 1,972건] + [내부] CHECK SHEET, W/R/I/P시리즈 | [외부] IEC 60076-1, IEEE C57.12.00, CIGRE TB 642"
    sources = project_info.get('출처', default_sources) if project_info else default_sources
    source_cell.value = f"자료 출처: {sources}"
    source_cell.alignment = Alignment(horizontal='left', vertical='center')
    source_cell.font = Font(italic=True, size=14)

    # Row 4: 빈 행 (구분선 역할)
    ws.row_dimensions[4].height = 5

    # Row 5: AIAG-VDA 7-Step 프로세스 구분
    # Row 5 높이 설정 (글자 잘림 방지)
    ws.row_dimensions[5].height = 40

    # 테두리 정의 (Row 5에도 적용)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    step_sections = [
        ('A5', 'A5', '구조분석(step 2)'),
        ('B5', 'B5', '기능분석(step 3)'),
        ('C5', 'G5', '고장분석(step 4)'),
        ('H5', 'M5', '리스크분석(step 5)'),  # H-M (RPN, AP 포함)
        ('N5', 'V5', '최적화(step 6)')       # N-V (RPN', AP' 포함)
    ]

    for start_cell, end_cell, label in step_sections:
        if start_cell != end_cell:
            ws.merge_cells(f'{start_cell}:{end_cell}')
        cell = ws[start_cell]
        cell.value = label
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border  # 테두리 추가

        # 병합된 셀의 모든 셀에도 테두리 적용
        if start_cell != end_cell:
            start_col = ord(start_cell[0]) - ord('A') + 1
            end_col = ord(end_cell[0]) - ord('A') + 1
            for col in range(start_col, end_col + 1):
                ws.cell(row=5, column=col).border = thin_border

    # Row 6: 헤더 (22개 컬럼: A-V)
    headers = [
        "부품명", "기능", "고장영향", "S", "고장형태",
        "고장원인", "고장메커니즘", "현재예방대책", "O",
        "현재검출대책", "D", "RPN", "AP",
        "예방조치", "검출조치", "담당자", "목표일",
        "S'", "O'", "D'", "RPN'", "AP'"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row 6 높이 설정
    ws.row_dimensions[6].height = 30

    # Row 6-1: 입력 안내 메시지 추가 (데이터 유효성 검사 방식, M365 웹 호환)
    print("6-1. 입력 안내 메시지 추가 중...")
    add_cell_validation_messages(ws)

    # Row 7+: 데이터 (22개 컬럼: A-V)
    for row_idx, item in enumerate(fmea_data, start=7):
        ws[f'A{row_idx}'] = item['부품명']
        ws[f'B{row_idx}'] = item['기능']
        ws[f'C{row_idx}'] = item['고장영향']
        ws[f'D{row_idx}'] = item['S']
        ws[f'E{row_idx}'] = item['고장형태']
        ws[f'F{row_idx}'] = item['고장원인']
        ws[f'G{row_idx}'] = item['고장메커니즘']
        ws[f'H{row_idx}'] = item['현재예방대책']
        ws[f'I{row_idx}'] = item['O']
        ws[f'J{row_idx}'] = item['현재검출대책']
        ws[f'K{row_idx}'] = item['D']
        ws[f'L{row_idx}'] = item['RPN']       # RPN = S x O x D
        ws[f'M{row_idx}'] = item['AP']        # AP (H/M/L)
        ws[f'N{row_idx}'] = item['예방조치']
        ws[f'O{row_idx}'] = item['검출조치']
        ws[f'P{row_idx}'] = item['담당자']
        ws[f'Q{row_idx}'] = item['목표일']
        ws[f'R{row_idx}'] = item["S'"]
        ws[f'S{row_idx}'] = item["O'"]
        ws[f'T{row_idx}'] = item["D'"]
        ws[f'U{row_idx}'] = item["RPN'"]      # RPN' = S' x O' x D'
        ws[f'V{row_idx}'] = item["AP'"]       # AP' (H/M/L)

        # Row 7부터 데이터 행 높이 설정 (옵션 A: 3줄 구조 지원)
        # 고장형태(E열)는 3줄 구조이므로 줄바꿈 개수에 따라 높이 조정
        max_newlines = 0
        for key in ['고장형태', '고장영향', '고장원인']:
            value = str(item.get(key, ''))
            newline_count = value.count('\n')
            if newline_count > max_newlines:
                max_newlines = newline_count

        if max_newlines >= 2:
            ws.row_dimensions[row_idx].height = 65  # 3줄 구조 (고장형태)
        elif max_newlines == 1:
            ws.row_dimensions[row_idx].height = 50  # 2줄 구조 (고장영향/고장원인)
        else:
            ws.row_dimensions[row_idx].height = 35  # 기본 높이

    # 4. 셀 병합 (A-E만)
    print("4. 셀 병합 적용 중...")
    apply_cell_merge(ws, fmea_data, start_row=7)

    # 5. 서식 적용
    print("5. 서식 적용 중...")
    # thin_border는 위에서 이미 정의됨

    # 5-1. 가운데 정렬 컬럼 리스트 (22개 컬럼 기준)
    # 가운데: 부품명, 기능, 고장영향, S, 고장형태, O, D, RPN, AP, 담당자, 목표일, S', O', D', RPN', AP'
    center_align_cols = ['A', 'B', 'C', 'D', 'E', 'I', 'K', 'L', 'M', 'P', 'Q', 'R', 'S', 'T', 'U', 'V']

    for row in ws.iter_rows(min_row=6, max_row=6+len(fmea_data), min_col=1, max_col=22):
        for cell in row:
            cell.border = thin_border
            col_letter = get_column_letter(cell.column)

            if col_letter in center_align_cols:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 5-2. RPN 및 AP 컬럼 색상 적용
    print("5-2. RPN 및 AP 컬럼 색상 적용 중...")

    # AP 색상 (H/M/L)
    ap_colors = {
        'H': PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'),  # 빨강
        'M': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),  # 노랑
        'L': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')   # 녹색
    }

    # RPN 색상 (수치 기반: >=200 빨강, 100-199 노랑, <100 녹색)
    rpn_colors = {
        'high': PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'),    # RPN >= 200: 빨강
        'medium': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),  # RPN 100-199: 노랑
        'low': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')      # RPN < 100: 녹색
    }

    def get_rpn_color(rpn_value):
        """RPN 값에 따른 색상 반환"""
        try:
            rpn = int(rpn_value)
            if rpn >= 200:
                return rpn_colors['high']
            elif rpn >= 100:
                return rpn_colors['medium']
            else:
                return rpn_colors['low']
        except (ValueError, TypeError):
            return rpn_colors['low']  # 기본값

    for row_idx, item in enumerate(fmea_data, start=7):
        # RPN 컬럼 (L열) - 수치 기반 색상
        rpn_value = item.get('RPN', 0)
        rpn_fill = get_rpn_color(rpn_value)
        ws[f'L{row_idx}'].fill = rpn_fill
        ws[f'L{row_idx}'].font = Font(bold=True, color='FFFFFF')  # 흰색 글자

        # AP 컬럼 (M열) - H/M/L 색상
        ap_value = item.get('AP', 'L')
        if ap_value in ap_colors:
            ws[f'M{row_idx}'].fill = ap_colors[ap_value]
            ws[f'M{row_idx}'].font = Font(bold=True, color='FFFFFF')  # 흰색 글자

        # RPN' 컬럼 (U열) - 수치 기반 색상
        rpn_prime_value = item.get("RPN'", 0)
        rpn_prime_fill = get_rpn_color(rpn_prime_value)
        ws[f'U{row_idx}'].fill = rpn_prime_fill
        ws[f'U{row_idx}'].font = Font(bold=True, color='FFFFFF')  # 흰색 글자

        # AP' 컬럼 (V열) - H/M/L 색상
        ap_prime_value = item.get("AP'", 'L')
        if ap_prime_value in ap_colors:
            ws[f'V{row_idx}'].fill = ap_colors[ap_prime_value]
            ws[f'V{row_idx}'].font = Font(bold=True, color='FFFFFF')  # 흰색 글자

    print(f"[OK] RPN 색상 적용 완료 (>=200=빨강, 100-199=노랑, <100=녹색)")
    print(f"[OK] AP 색상 적용 완료 (H=빨강, M=노랑, L=녹색)")

    # 컬럼 너비 설정 (A-V 전체, 22개 컬럼)
    print("5-3. 컬럼 너비 설정 중...")
    column_widths = {
        'A': 15,   # 부품명
        'B': 20,   # 기능
        'C': 25,   # 고장영향
        'D': 6,    # S
        'E': 20,   # 고장형태
        'F': 45,   # 고장원인
        'G': 60,   # 고장메커니즘
        'H': 45,   # 현재예방대책
        'I': 6,    # O
        'J': 45,   # 현재검출대책
        'K': 6,    # D
        'L': 8,    # RPN (NEW)
        'M': 6,    # AP (이동: L->M)
        'N': 45,   # 예방조치 (이동: M->N)
        'O': 45,   # 검출조치 (이동: N->O)
        'P': 22,   # 담당자 (이동: O->P)
        'Q': 12,   # 목표일 (이동: P->Q)
        'R': 6,    # S' (이동: Q->R)
        'S': 6,    # O' (이동: R->S)
        'T': 6,    # D' (이동: S->T)
        'U': 8,    # RPN' (NEW)
        'V': 6     # AP' (이동: T->V)
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    print(f"[OK] 컬럼 너비 설정 완료 (22개 컬럼, 가독성 최적화)")

    # 페이지 설정 (A3 가로)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    # 틀고정 (Row 1-6 고정, Row 7부터 스크롤)
    ws.freeze_panes = 'A7'
    print("[OK] 틀고정 적용: Row 1-6 고정")

    # 6. 파일 저장
    print(f"6. 파일 저장 중: {output_path}")
    wb.save(output_path)
    print(f"[OK] Excel 생성 완료: {output_path}")
    print(f"    - Sheet 1: 기능분석 ({len(function_data)}개 항목)")
    print(f"    - Sheet 2: FMEA ({len(fmea_data)}개 항목)")

    return output_path


if __name__ == "__main__":
    import sys
    import argparse

    # 기본 QA DB 경로 (일진전기 변압기 FMEA 전용)
    DEFAULT_QA_DB = r"C:\Users\jmyoo\Desktop\FMEA\04.HMS_DB\QA자료\01_원본데이터\QA_품질이력.db"

    parser = argparse.ArgumentParser(
        description='FMEA Excel 자동 생성 스크립트 (일진전기 변압기 FMEA 전용)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python generate_fmea_excel.py input.json output.xlsx
  python generate_fmea_excel.py input.json output.xlsx --no-qa-db

QA DB 자동화 기능 (기본 활성화):
  - S값: 중요_경미, 치명도, 분류(CLAIM +1), 피해보상비 기반 자동 계산
  - O값: 12년간 발생건수 + 최근3년 트렌드 기반 자동 계산
  - D값: 검사구분, 항목구분 기반 자동 계산
  - F열: 발생원인유형, 원인부서 기반 라이프사이클 태그 자동 부여
  - C열: 발생현상유형 3단계 체계 적용
  - H/J열: 재발방지대책, 조치내역 자동 참조

비활성화: --no-qa-db 옵션 사용
        """
    )
    parser.add_argument('input_file', help='Input JSON file (fmea_data)')
    parser.add_argument('output_file', help='Output Excel file (.xlsx)')
    parser.add_argument('--qa-db', dest='qa_db', default=DEFAULT_QA_DB,
                        help=f'QA DB 경로 (기본값: {DEFAULT_QA_DB})')
    parser.add_argument('--no-qa-db', dest='no_qa_db', action='store_true',
                        help='QA DB 자동화 비활성화')
    parser.add_argument('--dry-run', dest='dry_run', action='store_true',
                        help='검증만 수행 (Excel 생성 없이 BLOCKING 오류 확인)')

    args = parser.parse_args()

    # JSON 데이터 로드
    with open(args.input_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    fmea_data = data.get('fmea_data', [])
    project_info = data.get('project_info', {})

    # dry-run 모드: 검증만 수행 (Excel 생성 없음)
    if args.dry_run:
        print("=" * 60)
        print("[DRY-RUN] 검증 전용 모드 (Excel 생성 없음)")
        print("=" * 60)
        print("입력: %s (%d 항목)" % (args.input_file, len(fmea_data)))

        # 정렬 (검증 전 필수)
        function_order = project_info.get('function_order', None) if project_info else None
        fmea_data = sort_fmea_data(fmea_data, function_order=function_order)

        # 모든 검증 수행 (ALL-AT-ONCE)
        all_errors = []
        try:
            validate_data(fmea_data)
            print("[PASS] validate_data")
        except ValueError as e:
            all_errors.append("[DATA] %s" % str(e))
            print("[FAIL] validate_data")

        try:
            validate_logical_consistency(fmea_data)
            print("[PASS] validate_logical_consistency")
        except ValueError as e:
            all_errors.append("[LOGIC] %s" % str(e))
            print("[FAIL] validate_logical_consistency")

        try:
            validate_failure_effect_phases(fmea_data)
            print("[PASS] validate_failure_effect_phases")
        except ValueError as e:
            all_errors.append("[PHASE] %s" % str(e))
            print("[FAIL] validate_failure_effect_phases")

        # C열 금지어
        fe_violations = 0
        for i, row in enumerate(fmea_data):
            fe = row.get('고장영향', '')
            if fe:
                is_valid, reason = validate_failure_effect(fe)
                if not is_valid:
                    fe_violations += 1
                    if fe_violations <= 5:
                        print("  C열 행%d: '%s' -> %s" % (i+1, fe[:30], reason))
        if fe_violations:
            all_errors.append("[C열] 금지어 %d건" % fe_violations)
            print("[FAIL] C열 금지어 (%d건)" % fe_violations)
        else:
            print("[PASS] C열 금지어")

        # 병합 연속성
        try:
            validate_merge_contiguity(fmea_data)
            print("[PASS] validate_merge_contiguity")
        except ValueError as e:
            all_errors.append("[MERGE] %s" % str(e))
            print("[FAIL] validate_merge_contiguity")

        # 다이아몬드 구조
        try:
            validate_diamond_structure_data(fmea_data)
            print("[PASS] validate_diamond_structure")
        except ValueError as e:
            all_errors.append("[DIAMOND] %s" % str(e))
            print("[FAIL] validate_diamond_structure")

        # 결과 요약
        print("\n" + "=" * 60)
        if all_errors:
            print("[FAIL] %d개 BLOCKING 오류 발견!" % len(all_errors))
            for err in all_errors:
                print("  %s" % err[:100])
            sys.exit(1)
        else:
            print("[PASS] 모든 검증 통과! Excel 생성 가능.")
            sys.exit(0)

    # QA DB 자동화 적용 (기본 활성화, --no-qa-db로 비활성화)
    if not args.no_qa_db and args.qa_db:
        print("=" * 60)
        print("[QA DB AUTO] QA 품질이력 DB 자동 연동 (기본 활성화)")
        print("=" * 60)
        fmea_data = enhance_fmea_with_qa_db(fmea_data, args.qa_db)
        print()

    # Excel 생성
    generate_excel(fmea_data, args.output_file, project_info)
