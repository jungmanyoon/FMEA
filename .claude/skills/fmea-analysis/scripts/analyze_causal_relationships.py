# -*- coding: utf-8 -*-
"""
철심_FMEA.xlsx 인과관계 상세 분석
왜 80%가 실패했는지 근본 원인 파악

온톨로지:
    - references/function-effect-ontology.md에서 키워드 매핑 동적 로드
    - SSOT: 기능-고장영향 키워드 매핑은 온톨로지 파일에서 관리
"""

import openpyxl
import sys
import io
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 스크립트 디렉토리
script_dir = Path(__file__).parent


def load_function_effect_keywords() -> dict:
    """
    function-effect-ontology.md에서 기능-고장영향 키워드 매핑 로드

    Returns:
        {'자속': {'related_effects': [...], 'description': '...'}, ...}
    """
    ontology_path = script_dir.parent / "references" / "function-effect-ontology.md"

    result = {}

    if not ontology_path.exists():
        # 폴백: 기본값 반환
        return {
            '손실': {'related_effects': ['손실', '효율', '열', '온도'], 'description': '손실 증가 -> 열/온도 상승'},
            '자속': {'related_effects': ['전압', '변환', '손실', '효율'], 'description': '자속 전달 실패 -> 변환 불가'},
            '지지': {'related_effects': ['변형', '진동', '소음'], 'description': '지지 실패 -> 구조 변형'},
            '절연': {'related_effects': ['지락', '절연', '파괴'], 'description': '절연 실패 -> 지락'},
            '냉각': {'related_effects': ['과열', '온도', '열'], 'description': '냉각 실패 -> 과열'},
        }

    content = ontology_path.read_text(encoding='utf-8')

    # SECTION:FUNCTION_EFFECT_KEYWORDS 파싱
    in_section = False
    for line in content.split('\n'):
        if 'SECTION:FUNCTION_EFFECT_KEYWORDS' in line:
            in_section = True
            continue
        if in_section and line.startswith('## SECTION:'):
            break
        if in_section and ':' in line and not line.startswith('#'):
            parts = line.split(':', 1)
            if len(parts) == 2:
                key = parts[0].strip()
                values = [v.strip() for v in parts[1].split(',') if v.strip()]
                if key and values:
                    result[key] = {
                        'related_effects': values,
                        'description': f'{key} 실패 -> {", ".join(values[:3])}'
                    }

    return result


# 온톨로지에서 로드 (모듈 로드 시 1회)
_function_keywords = load_function_effect_keywords()

def expand_merged_cells(wb, sheet_name='FMEA'):
    """병합된 셀 복원"""
    ws = wb[sheet_name]
    headers = [ws.cell(6, col).value for col in range(1, 21)]

    data = []
    prev_values = {}

    for row_idx in range(7, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            val = ws.cell(row_idx, col_idx).value
            if val is None:
                row_data[header] = prev_values.get(header, None)
            else:
                row_data[header] = val
                prev_values[header] = val
        data.append(row_data)

    return data

def analyze_causal_chain(data):
    """인과관계 체인 상세 분석"""
    print("\n" + "="*100)
    print("인과관계 상세 분석 (Causal Chain Analysis)")
    print("="*100)

    # 기능별 그룹화
    function_groups = {}
    for i, row in enumerate(data, 1):
        func = row.get('기능', '')
        if func and func not in function_groups:
            function_groups[func] = []
        if func:
            function_groups[func].append((i, row))

    print(f"\n총 기능 개수: {len(function_groups)}개")

    # 각 기능별 분석
    for func_idx, (func, rows) in enumerate(function_groups.items(), 1):
        print(f"\n{'='*100}")
        print(f"[기능 {func_idx}] {func}")
        print(f"{'='*100}")
        print(f"항목 개수: {len(rows)}개")

        # 고장영향 추출
        effects = set()
        modes = set()
        causes = set()

        for row_num, row in rows:
            effect = row.get('고장영향', '')
            mode = row.get('고장형태', '')
            cause = row.get('고장원인', '')

            if effect:
                effects.add(effect)
            if mode:
                modes.add(mode)
            if cause:
                causes.add(cause)

        print(f"\n[구조]")
        print(f"  1개 기능 -> {len(effects)}개 고장영향 -> {len(modes)}개 고장형태 -> {len(causes)}개 고장원인")

        # 각 고장영향별 상세 출력
        for effect in effects:
            print(f"\n  [고장영향: {effect}]")

            # 인과관계 체크
            print(f"    기능-영향 인과관계 체크:")
            check_function_effect_causality(func, effect)

            # 이 영향에 속하는 형태들
            effect_modes = []
            for row_num, row in rows:
                if row.get('고장영향', '') == effect:
                    mode = row.get('고장형태', '')
                    cause = row.get('고장원인', '')
                    mech = row.get('고장메커니즘', '')
                    if mode:
                        effect_modes.append({
                            'row': row_num,
                            'mode': mode,
                            'cause': cause,
                            'mechanism': mech
                        })

            print(f"    관련 고장형태 {len(effect_modes)}개:")
            for item in effect_modes[:3]:  # 처음 3개만
                print(f"      - {item['mode']}")
                print(f"        원인: {item['cause']}")
                print(f"        메커니즘: {item['mechanism']}")

def check_function_effect_causality(func, effect):
    """기능-고장영향 인과관계 체크 (온톨로지에서 로드)"""

    # 기능-고장영향 키워드 매핑 (온톨로지에서 로드)
    function_keywords = _function_keywords

    # 기능에서 키워드 추출
    func_keyword = None
    for keyword, info in function_keywords.items():
        if keyword in func:
            func_keyword = keyword
            break

    if not func_keyword:
        print(f"      [WARN] 기능 '{func[:30]}...' - 알려진 키워드 없음")
        return

    # 고장영향에서 관련 키워드 찾기
    related = function_keywords[func_keyword]['related_effects']
    matched = any(kw in effect for kw in related)

    if matched:
        print(f"      [OK] 인과관계 논리적: {func[:30]}... -> {effect[:30]}...")
        print(f"         ({function_keywords[func_keyword]['description']})")
    else:
        print(f"      [NG] 인과관계 불명확: {func[:30]}... -> {effect[:30]}...")
        print(f"         예상 효과: {', '.join(related[:5])}")
        print(f"         실제 효과: {effect}")

def analyze_keyword_coverage(data):
    """키워드 커버리지 분석"""
    print("\n" + "="*100)
    print("키워드 커버리지 분석")
    print("="*100)

    # 현재 검증 로직 키워드
    current_keywords = {
        '자속': ['전압', '변환', '손실', '효율', '여자전류', '무부하'],
        '지지': ['변형', '진동', '소음', '정렬', '정밀도'],
        '절연': ['지락', '절연', '전기'],
        '냉각': ['과열', '온도', '열'],
    }

    # 실제 데이터에서 나타나는 패턴 수집
    function_effects_actual = {}
    for row in data:
        func = row.get('기능', '')
        effect = row.get('고장영향', '')

        if func and effect:
            if func not in function_effects_actual:
                function_effects_actual[func] = set()
            function_effects_actual[func].add(effect)

    print("\n[실제 데이터 패턴]")
    for func, effects in function_effects_actual.items():
        print(f"\n기능: {func}")
        print(f"  고장영향 {len(effects)}개:")
        for effect in list(effects)[:5]:
            print(f"    - {effect}")

            # 현재 키워드로 매칭되는지 체크
            matched = False
            for func_kw, effect_kws in current_keywords.items():
                if func_kw in func:
                    if any(ekw in effect for ekw in effect_kws):
                        matched = True
                        print(f"      [OK] 매칭됨 (키워드: {func_kw})")
                    else:
                        print(f"      [NG] 매칭 실패 (키워드: {func_kw}, 예상: {effect_kws})")
                    break

            if not matched:
                print(f"      [WARN] 기능 키워드 없음")

def main():
    filepath = 'c:/Users/jmyoo/.claude/skills/fmea-analysis/철심_FMEA.xlsx'

    print("="*100)
    print("철심_FMEA.xlsx 인과관계 근본 원인 분석")
    print("="*100)

    wb = openpyxl.load_workbook(filepath)
    data = expand_merged_cells(wb)

    print(f"\n총 데이터 행: {len(data)}개")

    # 1. 인과관계 체인 상세 분석
    analyze_causal_chain(data)

    # 2. 키워드 커버리지 분석
    analyze_keyword_coverage(data)

    print("\n" + "="*100)
    print("분석 완료")
    print("="*100)

if __name__ == "__main__":
    main()
