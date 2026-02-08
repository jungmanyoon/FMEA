# -*- coding: utf-8 -*-
"""
초고압 변압기 SOD 평가기준 Excel 생성 스크립트 v2
- CIGRE TB 642 통계 기반
- 변압기 특화 시험방법 반영
"""

import sys
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

# Windows cp949 인코딩 문제 해결
if sys.stdout:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# 워크북 생성
wb = openpyxl.Workbook()

# 스타일 정의
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_font = Font(bold=True, color='FFFFFF', size=11)
header_font_black = Font(bold=True, size=11)
title_font = Font(bold=True, size=16)
subtitle_font = Font(bold=True, size=12)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 색상 정의
red_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
orange_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
yellow_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
green_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
light_green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
blue_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
light_blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

def get_score_color(score):
    """점수에 따른 색상 반환"""
    if score >= 9: return red_fill
    elif score >= 7: return orange_fill
    elif score >= 5: return yellow_fill
    elif score >= 3: return green_fill
    else: return light_green_fill

def get_score_font(score):
    """점수에 따른 폰트 반환"""
    if score >= 5:
        return Font(bold=True, color='FFFFFF')
    else:
        return Font(bold=True, color='000000')

# ============ Sheet 1: 심각도(Severity) ============
ws1 = wb.active
ws1.title = '심각도(Severity)'

# 제목
ws1.merge_cells('A1:D1')
ws1['A1'] = '심각도(Severity) - 초고압 변압기 고장영향 기준'
ws1['A1'].font = title_font
ws1['A1'].alignment = center_align
ws1.row_dimensions[1].height = 30

# 부제목
ws1.merge_cells('A2:D2')
ws1['A2'] = '※ 고장 발생 시 변압기 및 전력계통에 미치는 영향 정도'
ws1['A2'].font = Font(italic=True, color='666666')
ws1['A2'].alignment = left_align

# 헤더
headers_s = ['S', '등급', '고장영향', '변압기 특화 기준']
col_widths_s = [6, 12, 20, 60]
for col, (header, width) in enumerate(zip(headers_s, col_widths_s), 1):
    cell = ws1.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = blue_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws1.column_dimensions[get_column_letter(col)].width = width

# 데이터
severity_data = [
    [10, '치명적', '화재/폭발, 광역정전', '절연유 화재, 인접설비 손상, 154kV 이상 계통 블랙아웃'],
    [9, '위험', '변압기 완파, 장기정전', '권선 소손, 철심 손상으로 폐기 수준, 복구 6개월 이상'],
    [8, '매우 높음', '주요부품 교체 필요', '부싱/OLTC/권선 교체 필요, 복구 1~3개월'],
    [7, '높음', '긴급 운전정지', '보호계전기 트립, 즉시 운전정지 후 점검 필요'],
    [6, '보통', '제한운전 필요', '정격 용량 감소 운전, 부하 제한 필요'],
    [5, '낮음', '경보 발생', 'DGA 경보, 온도 경보, 계획정지 후 점검 필요'],
    [4, '경미', '모니터링 강화', '이상 징후 발견, 감시 주기 단축 필요'],
    [3, '사소', '정기점검 시 발견', '정상 운전 가능, 다음 정기점검 시 조치'],
    [2, '매우 사소', '외관 불량', '도장 벗겨짐, 명판 오류, 기능에 영향 없음'],
    [1, '없음', '영향 없음', '고장 시에도 영향 없음'],
]

for row_idx, row_data in enumerate(severity_data, 5):
    ws1.row_dimensions[row_idx].height = 35
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align if col_idx <= 3 else left_align
        if col_idx == 1:  # S 열 색상
            cell.fill = get_score_color(value)
            cell.font = get_score_font(value)

# ============ Sheet 2: 발생도(Occurrence) ============
ws2 = wb.create_sheet('발생도(Occurrence)')

# 제목
ws2.merge_cells('A1:D1')
ws2['A1'] = '발생도(Occurrence) - CIGRE 통계 기반'
ws2['A1'].font = title_font
ws2['A1'].alignment = center_align
ws2.row_dimensions[1].height = 30

# 부제목
ws2.merge_cells('A2:D2')
ws2['A2'] = '※ CIGRE TB 642: 변압기 평균 고장률 0.5%/년(변전소), 1%/년(GSU)'
ws2['A2'].font = Font(italic=True, color='666666')
ws2['A2'].alignment = left_align

# 헤더
headers_o = ['O', '등급', '고장률', '변압기 특화 기준']
col_widths_o = [6, 12, 15, 55]
for col, (header, width) in enumerate(zip(headers_o, col_widths_o), 1):
    cell = ws2.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = blue_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(col)].width = width

# 데이터
occurrence_data = [
    [10, '극히 높음', '> 2%/년', '설계 결함, 과거 동일 고장 반복 발생'],
    [9, '매우 높음', '1~2%/년', '유사 설계에서 고장 이력 있음, GSU 700kV급 수준'],
    [8, '높음', '0.5~1%/년', '업계 평균 수준, 개선 필요'],
    [7, '다소 높음', '0.2~0.5%/년', '평균 이하지만 목표 미달'],
    [6, '보통', '0.1~0.2%/년', '목표 고장률 수준 (우수 제조사 기준)'],
    [5, '다소 낮음', '0.05~0.1%/년', '검증된 설계, 품질관리 우수'],
    [4, '낮음', '0.02~0.05%/년', '충분한 운전 실적, 안정적 설계'],
    [3, '매우 낮음', '0.01~0.02%/년', '다수 실적, 고장 사례 거의 없음'],
    [2, '희박', '< 0.01%/년', '표준화된 설계, 수십 년 무고장 실적'],
    [1, '거의 불가능', '0', '설계적으로 원천 차단, 발생 불가'],
]

for row_idx, row_data in enumerate(occurrence_data, 5):
    ws2.row_dimensions[row_idx].height = 30
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align if col_idx <= 3 else left_align
        if col_idx == 1:  # O 열 색상
            cell.fill = get_score_color(value)
            cell.font = get_score_font(value)

# O 조정 기준 추가
ws2.merge_cells('A16:D16')
ws2['A16'] = '> 발생도 산정 보조 기준'
ws2['A16'].font = subtitle_font
ws2['A16'].alignment = left_align

adjust_headers = ['조건', 'O 조정', '비고', '']
for col, header in enumerate(adjust_headers, 1):
    cell = ws2.cell(row=17, column=col, value=header)
    cell.font = header_font_black
    cell.fill = light_blue_fill
    cell.alignment = center_align
    cell.border = thin_border

adjust_data = [
    ['신규 설계 (전압등급 변경, 신규 부품 적용)', '+2', '검증 이력 없음'],
    ['용량 증대 설계 (기존 대비 20% 이상)', '+1', '열적/기계적 스트레스 증가'],
    ['검증된 표준 설계 반복 생산', '-1', '실적 보유'],
    ['형식시험 통과 이력 있음', '-1', '설계 검증 완료'],
    ['유사 설계 현장 고장 이력 있음', '+2', '리스크 증가'],
]

for row_idx, row_data in enumerate(adjust_data, 18):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align if col_idx == 2 else left_align

ws2.merge_cells('C18:D18')
ws2.merge_cells('C19:D19')
ws2.merge_cells('C20:D20')
ws2.merge_cells('C21:D21')
ws2.merge_cells('C22:D22')

# ============ Sheet 3: 검출도(Detection) ============
ws3 = wb.create_sheet('검출도(Detection)')

# 제목
ws3.merge_cells('A1:D1')
ws3['A1'] = '검출도(Detection) - 변압기 시험방법 기준'
ws3['A1'].font = title_font
ws3['A1'].alignment = center_align
ws3.row_dimensions[1].height = 30

# 부제목
ws3.merge_cells('A2:D2')
ws3['A2'] = '※ DGA는 잠재 고장의 약 70% 검출 가능 (IEC 60599)'
ws3['A2'].font = Font(italic=True, color='666666')
ws3['A2'].alignment = left_align

# 헤더
headers_d = ['D', '등급', '검출 가능성', '검출 방법 (변압기 특화)']
col_widths_d = [6, 12, 20, 55]
for col, (header, width) in enumerate(zip(headers_d, col_widths_d), 1):
    cell = ws3.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = blue_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws3.column_dimensions[get_column_letter(col)].width = width

# 데이터
detection_data = [
    [10, '검출 불가', '검출 수단 없음', '운전 중 고장 발생 시에만 인지, 예방 검출 불가'],
    [9, '극히 어려움', '특수 분석 필요', '고장 후 정밀 분석으로만 판별 (해체 검사)'],
    [8, '매우 어려움', '온라인 모니터링 한계', 'DGA 트렌드 분석 필요, 즉시 검출 어려움'],
    [7, '어려움', '정밀 진단시험 필요', 'PD 측정, FRA 분석 등 특수장비 필요'],
    [6, '보통', '정기 진단시험', 'DGA 샘플링, 절연유 분석 (주기적)'],
    [5, '다소 용이', 'FAT 시험 검출', '내전압시험, 유도시험, 부분방전 측정'],
    [4, '용이', '루틴시험 검출', '변압비, 권선저항, 절연저항 측정'],
    [3, '매우 용이', '공정검사 검출', '조립 중 치수검사, 외관검사, QC 점검'],
    [2, '높음', '수입검사 검출', '입고 시 검사로 불량 검출'],
    [1, '확실', '자동 검출', '설계적 Poka-Yoke, 자동 검사 시스템'],
]

for row_idx, row_data in enumerate(detection_data, 5):
    ws3.row_dimensions[row_idx].height = 30
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align if col_idx <= 3 else left_align
        if col_idx == 1:  # D 열 색상
            cell.fill = get_score_color(value)
            cell.font = get_score_font(value)

# D 보조 기준 추가
ws3.merge_cells('A16:D16')
ws3['A16'] = '> 검출도 산정 보조 기준 (시험방법별)'
ws3['A16'].font = subtitle_font
ws3['A16'].alignment = left_align

detect_headers = ['검출 방법', '기본 D', '비고', '']
for col, header in enumerate(detect_headers, 1):
    cell = ws3.cell(row=17, column=col, value=header)
    cell.font = header_font_black
    cell.fill = light_blue_fill
    cell.alignment = center_align
    cell.border = thin_border

detect_data = [
    ['온라인 DGA 모니터링', '5~6', '실시간 가스 분석'],
    ['오프라인 DGA 샘플링 (월 1회)', '6~7', '주기적 분석'],
    ['PD 측정 (UHF/AE)', '6~7', '부분방전 검출'],
    ['FRA (주파수응답분석)', '7', '권선 변형 검출'],
    ['FAT 내전압/유도시험', '4~5', '출하 전 절연 검증'],
    ['변압비/권선저항 측정', '3~4', '루틴시험'],
    ['절연저항 측정', '4', '루틴시험'],
    ['육안검사/치수검사', '2~3', '공정검사'],
]

for row_idx, row_data in enumerate(detect_data, 18):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align if col_idx == 2 else left_align

ws3.merge_cells('C18:D18')
ws3.merge_cells('C19:D19')
ws3.merge_cells('C20:D20')
ws3.merge_cells('C21:D21')
ws3.merge_cells('C22:D22')
ws3.merge_cells('C23:D23')
ws3.merge_cells('C24:D24')
ws3.merge_cells('C25:D25')

# ============ Sheet 4: AP 기준 ============
ws4 = wb.create_sheet('AP기준')

# 제목
ws4.merge_cells('A1:D1')
ws4['A1'] = 'AP (Action Priority) 산정 기준'
ws4['A1'].font = title_font
ws4['A1'].alignment = center_align
ws4.row_dimensions[1].height = 30

# 헤더
headers_ap = ['AP', '조건', '조치', '']
col_widths_ap = [8, 40, 40, 10]
for col, (header, width) in enumerate(zip(headers_ap, col_widths_ap), 1):
    cell = ws4.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = blue_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws4.column_dimensions[get_column_letter(col)].width = width

# 데이터
ap_data = [
    ['H (High)', 'S>=8 또는 SxO>=40 또는 RPN>=200', '즉시 설계 개선 필요', red_fill],
    ['M (Medium)', 'S>=5 또는 SxO>=20 또는 RPN>=100', '개선 검토 필요', yellow_fill],
    ['L (Low)', '그 외', '현 관리 수준 유지', light_green_fill],
]

for row_idx, (ap, condition, action, color) in enumerate(ap_data, 4):
    ws4.row_dimensions[row_idx].height = 40

    cell1 = ws4.cell(row=row_idx, column=1, value=ap)
    cell1.border = thin_border
    cell1.alignment = center_align
    cell1.fill = color
    cell1.font = Font(bold=True, color='FFFFFF' if color == red_fill else '000000')

    cell2 = ws4.cell(row=row_idx, column=2, value=condition)
    cell2.border = thin_border
    cell2.alignment = center_align

    cell3 = ws4.cell(row=row_idx, column=3, value=action)
    cell3.border = thin_border
    cell3.alignment = center_align

ws4.merge_cells('C4:D4')
ws4.merge_cells('C5:D5')
ws4.merge_cells('C6:D6')

# RPN 참고
ws4.merge_cells('A8:D8')
ws4['A8'] = '> RPN = S x O x D (1 ~ 1,000)'
ws4['A8'].font = subtitle_font

ws4.merge_cells('A9:D9')
ws4['A9'] = '※ RPN이 높을수록 위험 우선순위 높음. 단, S가 높은 항목은 RPN과 무관하게 우선 관리'
ws4['A9'].font = Font(italic=True, color='666666')

# ============ Sheet 5: 적용예시 ============
ws5 = wb.create_sheet('적용예시')

# 제목
ws5.merge_cells('A1:H1')
ws5['A1'] = '적용 예시: 권선 절연지 FMEA'
ws5['A1'].font = title_font
ws5['A1'].alignment = center_align
ws5.row_dimensions[1].height = 30

# 헤더
headers_ex = ['부품', '고장형태', '고장영향', 'S', 'O', 'D', 'RPN', 'AP']
col_widths_ex = [12, 10, 40, 5, 5, 5, 8, 8]
for col, (header, width) in enumerate(zip(headers_ex, col_widths_ex), 1):
    cell = ws5.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = blue_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws5.column_dimensions[get_column_letter(col)].width = width

# 데이터
example_data = [
    ['권선 절연지', '절연파괴', '층간단락 -> 권선 소손 -> 변압기 완파', 9, 3, 7, 189, 'M'],
    ['권선 절연지', '열화', '절연수명 단축 -> 제한운전', 6, 4, 6, 144, 'M'],
    ['권선 절연지', '크랙', 'PD 발생 -> DGA 경보', 5, 3, 5, 75, 'L'],
]

for row_idx, row_data in enumerate(example_data, 4):
    ws5.row_dimensions[row_idx].height = 35
    for col_idx, value in enumerate(row_data, 1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align

        # S 열 색상
        if col_idx == 4:
            cell.fill = get_score_color(value)
            cell.font = get_score_font(value)
        # AP 열 색상
        elif col_idx == 8:
            if value == 'H':
                cell.fill = red_fill
                cell.font = Font(bold=True, color='FFFFFF')
            elif value == 'M':
                cell.fill = yellow_fill
                cell.font = Font(bold=True)
            else:
                cell.fill = light_green_fill
                cell.font = Font(bold=True)

# S/O/D 산정 근거 추가
ws5.merge_cells('A8:H8')
ws5['A8'] = '> S/O/D 산정 근거'
ws5['A8'].font = subtitle_font

basis_data = [
    ['고장형태', 'S 근거', 'O 근거', 'D 근거'],
    ['절연파괴', 'S=9: 권선 소손 시 변압기 완파 수준', 'O=3: 표준 설계, 품질관리 우수', 'D=7: PD 측정으로 사전 검출 가능'],
    ['열화', 'S=6: 제한운전 필요, 긴급 정지 불필요', 'O=4: 장기 운전 시 발생 가능', 'D=6: DGA로 열화 가스 검출'],
    ['크랙', 'S=5: 경보 발생, 계획정지 후 점검', 'O=3: 제작 품질관리로 예방', 'D=5: FAT PD 측정으로 검출'],
]

for row_idx, row_data in enumerate(basis_data, 9):
    for col_idx, value in enumerate(row_data, 1):
        if col_idx == 1:
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        else:
            # 2열씩 병합
            ws5.merge_cells(start_row=row_idx, start_column=(col_idx-1)*2, end_row=row_idx, end_column=(col_idx-1)*2+1)
            cell = ws5.cell(row=row_idx, column=(col_idx-1)*2, value=value)

        cell.border = thin_border
        cell.alignment = left_align
        if row_idx == 9:
            cell.font = header_font_black
            cell.fill = light_blue_fill
            cell.alignment = center_align

# 저장
script_dir = os.path.dirname(os.path.abspath(__file__))
output_path = os.path.join(os.path.dirname(script_dir), '초고압변압기_SOD평가기준_v2.xlsx')
wb.save(output_path)
print(f'Excel 파일 생성 완료: {output_path}')
print('\n시트 구성:')
print('1. 심각도(Severity) - 변압기 고장영향 기준 (10단계)')
print('2. 발생도(Occurrence) - CIGRE 통계 기반 고장률 (10단계)')
print('3. 검출도(Detection) - 변압기 시험방법 기준 (10단계)')
print('4. AP기준 - Action Priority 산정 기준 (H/M/L)')
print('5. 적용예시 - 권선 절연지 FMEA 예시 + S/O/D 산정 근거')
