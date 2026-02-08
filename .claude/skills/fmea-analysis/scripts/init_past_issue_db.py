#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Past Issue Database Initializer (일진전기 초고압 변압기 과거 문제 DB)

This script initializes a structured database for tracking past issues
in transformer FMEA, enabling 60% content reuse target.

Usage:
    python init_past_issue_db.py [output_filename.xlsx]

Requirements:
    pip install openpyxl

Features:
    - Component-based issue tracking
    - Failure mode classification
    - Root cause analysis records
    - Noise factor documentation
    - Reusability scoring

Author: Claude (FMEA Analysis Skill)
Version: 2.0
Date: 2025-11-17
"""

import sys
import io
import argparse
from pathlib import Path
from datetime import datetime

# Windows cp949 인코딩 문제 해결
if sys.stdout:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("ERROR: openpyxl not installed")
    print("Install with: pip install openpyxl")
    sys.exit(1)


def create_db_structure(wb):
    """Create past issue database structure"""

    # Sheet 1: Database Overview
    ws_overview = wb.active
    ws_overview.title = "DB_Overview"

    overview_data = [
        ["일진전기 과거 문제 DB (Past Issue Database)", ""],
        ["생성 일자", datetime.now().strftime("%Y-%m-%d")],
        ["버전", "1.0"],
        ["목표 재활용률", "60%"],
        ["", ""],
        ["사용 방법", ""],
        ["1단계", "부품별 시트에서 과거 문제 검색"],
        ["2단계", "유사 고장 형태 확인"],
        ["3단계", "원인/메커니즘 재활용"],
        ["4단계", "대책 및 검증 방법 참조"],
        ["", ""],
        ["Phase 1 (초기 6개월~1년)", "핵심 부품 50개 이상 등록"],
        ["Phase 2 (1~2년)", "200개 이상 확장, 노이즈 인자 추가"],
        ["Phase 3 (2년 이후)", "AI 분석, 예측 모델 적용"],
    ]

    for row_idx, row_data in enumerate(overview_data, 1):
        ws_overview.cell(row=row_idx, column=1, value=row_data[0])
        if len(row_data) > 1:
            ws_overview.cell(row=row_idx, column=2, value=row_data[1])

    # Format header
    ws_overview.cell(1, 1).font = Font(bold=True, size=14)
    ws_overview.cell(1, 1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_overview.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFF")

    # Sheet 2: Master Issue List
    ws_master = wb.create_sheet(title="Master_Issue_List")

    master_headers = [
        "Issue ID",
        "등록일",
        "부품명 (Component)",
        "부품 계층 (Level)",
        "고장 형태 (Failure Mode)",
        "고장 원인 (Failure Cause)",
        "고장 메커니즘 (Mechanism)",
        "심각도 (S)",
        "발생도 (O)",
        "검출도 (D)",
        "RPN",
        "발생 프로젝트",
        "발생 시기 (공정)",
        "노이즈 인자",
        "대책 내용",
        "검증 방법",
        "재활용 횟수",
        "재활용 가능 여부",
        "비고",
    ]

    for col_idx, header in enumerate(master_headers, 1):
        cell = ws_master.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_master.column_dimensions[get_column_letter(col_idx)].width = 15

    # Sheet 3: Component-Based Categorization
    component_categories = [
        ("권선_Winding", ["권선 (Winding)", "절연지", "도체", "압착 구조"]),
        ("철심_Core", ["철심 (Core)", "적층 구조", "클램프", "절연 코팅"]),
        ("절연유_Oil", ["절연유 (Insulating Oil)", "DGA 가스", "수분 함량", "산가"]),
        ("탱크_Tank", ["탱크 (Tank)", "용접부", "가스켓", "방청"]),
        ("부싱_Bushing", ["부싱 (Bushing)", "절연체", "도체 연결", "실링"]),
        ("냉각_Cooling", ["냉각 장치", "펌프", "라디에이터", "팬"]),
        ("탭절환_OLTC", ["탭 절환기 (OLTC)", "접점", "구동부", "절연유"]),
        ("감시_Monitoring", ["감시 장치", "온도계", "압력계", "DGA 센서"]),
    ]

    for sheet_name, example_components in component_categories:
        ws = wb.create_sheet(title=sheet_name)

        # Headers same as master list
        for col_idx, header in enumerate(master_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Add example components in description
        ws.cell(row=2, column=1, value=f"예시 부품: {', '.join(example_components)}")
        ws.merge_cells('A2:F2')

    # Sheet: Noise Factors Reference
    ws_noise = wb.create_sheet(title="Noise_Factors")

    noise_data = [
        ["노이즈 인자 분류", "세부 항목", "예시"],
        ["1. 외부 환경", "온도", "-40°C ~ +50°C"],
        ["1. 외부 환경", "습도", "0% ~ 100% RH"],
        ["1. 외부 환경", "낙뢰", "낙뢰 서지"],
        ["1. 외부 환경", "염해", "해안 지역 염분"],
        ["2. 계통/고객 사용", "과부하", "정격 대비 120%"],
        ["2. 계통/고객 사용", "단락 사고", "계통 단락 전류"],
        ["2. 계통/고객 사용", "전압 변동", "±10% 전압 변동"],
        ["3. 시스템 상호작용", "고조파", "3차, 5차 고조파"],
        ["3. 시스템 상호작용", "공진", "계통 공진"],
        ["4. 시간 경과 변화", "절연 열화", "30년 장기 운전"],
        ["4. 시간 경과 변화", "기계적 마모", "접점 마모"],
        ["5. 부품 간 변동", "제작 공차", "±5% 치수 공차"],
        ["5. 부품 간 변동", "재료 특성 편차", "절연지 밀도 편차"],
    ]

    for row_idx, row_data in enumerate(noise_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_noise.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
                cell.font = Font(bold=True, size=11, color="FFFFFF")

    return wb


def main():
    parser = argparse.ArgumentParser(
        description='Initialize Past Issue Database for transformer FMEA'
    )
    parser.add_argument(
        'output',
        nargs='?',
        default='Past_Issue_DB.xlsx',
        help='Output filename (default: Past_Issue_DB.xlsx)'
    )

    args = parser.parse_args()

    print(f"[INFO] Initializing Past Issue Database...")

    # Create workbook
    wb = Workbook()
    wb = create_db_structure(wb)

    # Save
    output_path = Path(args.output)
    wb.save(output_path)

    print(f"[OK] Past Issue DB created: {output_path.absolute()}")
    print(f"   Sheets: {len(wb.sheetnames)}")
    print(f"\nDatabase Structure:")
    print(f"  1. DB_Overview - 사용 방법 및 목표")
    print(f"  2. Master_Issue_List - 전체 문제 통합 목록")
    print(f"  3-10. Component sheets - 부품별 분류 (권선, 철심, 절연유 등)")
    print(f"  11. Noise_Factors - 노이즈 인자 참조표")
    print(f"\nPhase 1 Goal: 핵심 부품 50개 이상 등록 (6개월~1년)")
    print(f"Target Reuse Rate: 60%")


if __name__ == "__main__":
    main()
