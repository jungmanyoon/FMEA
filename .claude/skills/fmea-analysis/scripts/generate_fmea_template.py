#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FMEA Excel Template Generator (일진전기 초고압 변압기 특화)

This script generates an Excel (.xlsx) template for FMEA analysis
optimized for Iljin Electric high-voltage transformers (154kV~765kV).

Usage:
    python generate_fmea_template.py [output_filename.xlsx] [--method hyundai|aiag-vda]

Requirements:
    pip install openpyxl

Author: Claude (FMEA Analysis Skill)
Version: 2.0
Date: 2025-11-17
"""

import sys
import io
import argparse
from pathlib import Path

# Windows cp949 인코딩 문제 해결
if sys.stdout:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed")
    print("Install with: pip install openpyxl")
    sys.exit(1)


def create_hyundai_9step_template(wb):
    """Create Hyundai 9-step method template sheets"""

    # Sheet 1: Project Info
    ws_info = wb.active
    ws_info.title = "0.프로젝트 정보"

    headers_info = [
        ["항목", "내용"],
        ["제품명", "154kV 주상변압기"],
        ["프로젝트 코드", ""],
        ["작성자", ""],
        ["작성 팀", ""],
        ["작성 일자", ""],
        ["검토자", ""],
        ["승인자", ""],
        ["FMEA 방법", "현대차 9단계 순차 작성"],
        ["기준 표준", "AIAG-VDA 2019"],
    ]

    for row_idx, row_data in enumerate(headers_info, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, size=12, color="FFFFFF")

    # Sheet 2-9: Step sheets
    step_sheets = [
        ("1.범위정의", ["항목", "내용"]),
        ("2.구조분석", ["Part No.", "부품명", "수량", "기능", "계층"]),
        ("3.초점요소", ["Focus Element", "선정 사유", "우선순위"]),
        ("4.Boundary", ["Input", "Output", "Interface"]),
        ("5.기능분류", ["기능", "분류", "고장영향", "심각도(S)"]),
        ("6.과거이력노이즈", ["부품명", "과거 문제", "노이즈 인자", "발생 빈도"]),
        ("7.고장분석", ["고장형태", "고장원인", "고장메커니즘", "S", "O", "D", "RPN", "조치사항"]),
        ("8.대책수립", ["조치 항목", "담당자", "완료 목표일", "진행 상태"]),
        ("9.결과확인", ["항목", "개선 전", "개선 후", "개선율(%)"]),
    ]

    for sheet_name, columns in step_sheets:
        ws = wb.create_sheet(title=sheet_name)
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    return wb


def create_aiag_vda_template(wb):
    """Create AIAG-VDA 7-step method template sheets"""

    ws_info = wb.active
    ws_info.title = "0.Project_Planning"

    headers_info = [
        ["Item", "Content"],
        ["Product Name", "Power Transformer"],
        ["Project Code", ""],
        ["Author", ""],
        ["Team", ""],
        ["Date", ""],
        ["Reviewer", ""],
        ["Approver", ""],
        ["FMEA Method", "AIAG-VDA 7-Step"],
        ["Standard", "AIAG-VDA 2019"],
    ]

    for row_idx, row_data in enumerate(headers_info, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, size=12, color="FFFFFF")

    step_sheets = [
        ("1.Planning", ["Item", "Description", "Responsibility"]),
        ("2.Structure", ["Item", "Function", "Requirements", "Interfaces"]),
        ("3.Function", ["Function", "Failure Mode", "Failure Effect", "S"]),
        ("4.Failure", ["Failure Mode", "Failure Cause", "O", "Current Controls", "D"]),
        ("5.Risk", ["Failure", "S", "O", "D", "AP", "Recommended Actions"]),
        ("6.Optimization", ["Action", "Responsibility", "Target Date", "Status"]),
        ("7.Documentation", ["Item", "Before", "After", "Improvement%"]),
    ]

    for sheet_name, columns in step_sheets:
        ws = wb.create_sheet(title=sheet_name)
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    return wb


def main():
    parser = argparse.ArgumentParser(
        description='Generate FMEA Excel template for Iljin Electric transformers'
    )
    parser.add_argument(
        'output',
        nargs='?',
        default='FMEA_Template.xlsx',
        help='Output filename (default: FMEA_Template.xlsx)'
    )
    parser.add_argument(
        '--method',
        choices=['hyundai', 'aiag-vda'],
        default='hyundai',
        help='FMEA method: hyundai (9-step) or aiag-vda (7-step) [default: hyundai]'
    )

    args = parser.parse_args()

    # Create workbook
    wb = Workbook()

    # Generate template based on method
    if args.method == 'hyundai':
        print(f"[INFO] Creating Hyundai 9-step method template...")
        wb = create_hyundai_9step_template(wb)
    else:
        print(f"[INFO] Creating AIAG-VDA 7-step method template...")
        wb = create_aiag_vda_template(wb)

    # Save workbook
    output_path = Path(args.output)
    wb.save(output_path)

    print(f"[OK] FMEA template created: {output_path.absolute()}")
    print(f"   Method: {args.method}")
    print(f"   Sheets: {len(wb.sheetnames)}")
    print(f"\nNext steps:")
    print(f"  1. Open {args.output} in Excel")
    print(f"  2. Fill in project information (Sheet 0)")
    print(f"  3. Follow step-by-step sheets in order")
    print(f"  4. For Hyundai method: Complete Steps 3-4-5 before clicking 'STEP4 button'")


if __name__ == "__main__":
    main()
