# -*- coding: utf-8 -*-
"""
철심_FMEA.xlsx 전체 검증 스크립트
현재 자동 검증 + 누락된 내용 검증 모두 수행

온톨로지:
    - references/function-effect-ontology.md에서 키워드 매핑 동적 로드
    - SSOT: 기능-고장영향 키워드 매핑은 온톨로지 파일에서 관리
"""

import openpyxl
import sys
import io
import re
from collections import defaultdict
from pathlib import Path

# UTF-8 출력 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 스크립트 디렉토리
script_dir = Path(__file__).parent


def load_function_effect_keywords() -> dict:
    """
    function-effect-ontology.md에서 기능-고장영향 키워드 매핑 로드

    Returns:
        {'자속': ['전압', '변환', ...], '지지': ['변형', ...], ...}
    """
    ontology_path = script_dir.parent / "references" / "function-effect-ontology.md"

    result = {}

    if not ontology_path.exists():
        # 폴백: 기본값 반환
        return {
            '자속': ['전압', '변환', '손실', '효율', '여자전류', '무부하'],
            '지지': ['변형', '진동', '소음', '정렬', '정밀도'],
            '절연': ['지락', '절연', '전기'],
            '냉각': ['과열', '온도', '열'],
        }

    content = ontology_path.read_text(encoding='utf-8')

    # SECTION:FUNCTION_EFFECT_KEYWORDS 파싱
    in_section = False
    for line in content.split('\n'):
        if 'SECTION:FUNCTION_EFFECT_KEYWORDS' in line:
            in_section = True
            continue
        if in_section and line.startswith('## SECTION:'):
            break
        if in_section and ':' in line and not line.startswith('#'):
            parts = line.split(':', 1)
            if len(parts) == 2:
                key = parts[0].strip()
                values = [v.strip() for v in parts[1].split(',') if v.strip()]
                if key and values:
                    result[key] = values

    return result


# 온톨로지에서 로드 (모듈 로드 시 1회)
_function_effect_keywords = load_function_effect_keywords()

def load_fmea_data(filepath):
    """Excel 파일에서 FMEA 데이터 로드"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['FMEA']

    # Header from Row 6
    headers = []
    for col in range(1, 21):  # A-T (20 columns)
        cell = ws.cell(row=6, column=col)
        headers.append(cell.value)

    # Data from Row 7 onwards
    data = []
    for row_idx in range(7, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            row_data[header] = val
        data.append(row_data)

    return data, headers

def expand_merged_cells(data):
    """병합된 셀 값을 복원 (NaN을 이전 값으로 채움)"""
    expanded = []
    prev_values = {}

    for row in data:
        new_row = {}
        for key, val in row.items():
            if val is None:
                new_row[key] = prev_values.get(key, None)
            else:
                new_row[key] = val
                prev_values[key] = val
        expanded.append(new_row)

    return expanded

def validate_causal_relationships(data):
    """인과관계 검증 (기능-영향-형태-원인)"""
    print("\n" + "="*80)
    print("1. 인과관계 검증 (Causal Relationship Validation)")
    print("="*80)

    issues = []

    # 기능-고장영향 키워드 매핑 (온톨로지에서 로드)
    function_effect_keywords = _function_effect_keywords

    for i, row in enumerate(data, 1):
        func = row.get('기능', '')
        effect = row.get('고장영향', '')
        mode = row.get('고장형태', '')
        cause = row.get('고장원인', '')

        if not all([func, effect, mode, cause]):
            continue

        # 기능-고장영향 인과관계 검증
        func_matched = False
        for func_kw, effect_kws in function_effect_keywords.items():
            if func_kw in func:
                if any(ekw in effect for ekw in effect_kws):
                    func_matched = True
                    break

        if not func_matched:
            issues.append({
                'row': i,
                'type': '기능-고장영향 인과관계',
                'severity': 'WARNING',
                'detail': f"기능 '{func[:30]}...' -> 고장영향 '{effect[:30]}...' 연관성 약함"
            })

    if issues:
        print(f"\n[경고] 인과관계 이슈 발견: {len(issues)}건")
        for issue in issues[:10]:
            print(f"  Row {issue['row']}: [{issue['type']}] {issue['detail']}")
        if len(issues) > 10:
            print(f"  ... 외 {len(issues) - 10}건")
    else:
        print("[OK] 인과관계 검증 통과")

    return issues

def validate_risk_ratings(data):
    """리스크 등급 검증 (S/O/D 범위, AP 계산)"""
    print("\n" + "="*80)
    print("2. 리스크 분석 검증 (Risk Rating Validation)")
    print("="*80)

    issues = []

    for i, row in enumerate(data, 1):
        S = row.get('S')
        O = row.get('O')
        D = row.get('D')
        AP = row.get('AP')

        # S/O/D 범위 검증 (1-10)
        for col_name, val in [('S', S), ('O', O), ('D', D)]:
            if val is not None:
                try:
                    int_val = int(val)
                    if not (1 <= int_val <= 10):
                        issues.append({
                            'row': i,
                            'type': f'{col_name}값 범위 오류',
                            'severity': 'ERROR',
                            'detail': f"{col_name}={val} (1-10 범위 초과)"
                        })
                except (ValueError, TypeError):
                    issues.append({
                        'row': i,
                        'type': f'{col_name}값 형식 오류',
                        'severity': 'ERROR',
                        'detail': f"{col_name}={val} (숫자가 아님)"
                    })

        # AP 계산 검증 (AIAG-VDA 기준)
        if all(v is not None for v in [S, O, D, AP]):
            try:
                S_int = int(S)
                O_int = int(O)
                D_int = int(D)

                # AIAG-VDA AP 계산 규칙
                if S_int >= 9 or (S_int >= 7 and O_int >= 4) or (S_int >= 4 and O_int >= 7):
                    expected_AP = 'H'
                elif S_int <= 3 and O_int <= 3 and D_int <= 3:
                    expected_AP = 'L'
                else:
                    expected_AP = 'M'

                if AP != expected_AP:
                    issues.append({
                        'row': i,
                        'type': 'AP 계산 오류',
                        'severity': 'WARNING',
                        'detail': f"AP={AP} (예상: {expected_AP}, S={S_int}, O={O_int}, D={D_int})"
                    })
            except (ValueError, TypeError):
                pass

    if issues:
        print(f"\n[오류] 리스크 분석 이슈 발견: {len(issues)}건")
        errors = [x for x in issues if x['severity'] == 'ERROR']
        warnings = [x for x in issues if x['severity'] == 'WARNING']

        if errors:
            print(f"\n  [ERROR] 오류 ({len(errors)}건):")
            for issue in errors[:5]:
                print(f"    Row {issue['row']}: [{issue['type']}] {issue['detail']}")
            if len(errors) > 5:
                print(f"    ... 외 {len(errors) - 5}건")

        if warnings:
            print(f"\n  [WARN] 경고 ({len(warnings)}건):")
            for issue in warnings[:5]:
                print(f"    Row {issue['row']}: [{issue['type']}] {issue['detail']}")
            if len(warnings) > 5:
                print(f"    ... 외 {len(warnings) - 5}건")
    else:
        print("[OK] 리스크 분석 검증 통과")

    return issues

def validate_improvement_effectiveness(data):
    """개선조치 효과 검증"""
    print("\n" + "="*80)
    print("3. 개선조치 효과 검증 (Improvement Action Validation)")
    print("="*80)

    issues = []

    for i, row in enumerate(data, 1):
        prevention = row.get('예방조치', '')
        detection = row.get('검출조치', '')
        O = row.get('O')
        D = row.get('D')

        # 예방조치가 있으면 구체적이어야 함
        if prevention:
            if len(prevention) < 5:
                issues.append({
                    'row': i,
                    'type': '예방조치 구체성',
                    'severity': 'WARNING',
                    'detail': f"예방조치가 너무 짧음: '{prevention}'"
                })

        # 검출조치가 있으면 구체적이어야 함
        if detection:
            if len(detection) < 5:
                issues.append({
                    'row': i,
                    'type': '검출조치 구체성',
                    'severity': 'WARNING',
                    'detail': f"검출조치가 너무 짧음: '{detection}'"
                })

    if issues:
        print(f"\n[경고] 개선조치 이슈 발견: {len(issues)}건")
        for issue in issues[:10]:
            print(f"  Row {issue['row']}: [{issue['type']}] {issue['detail']}")
        if len(issues) > 10:
            print(f"  ... 외 {len(issues) - 10}건")
    else:
        print("[OK] 개선조치 검증 통과")

    return issues

def validate_post_action_risk(data):
    """조치 후 리스크 검증 (S'<=S, O'<O, D'<D, AP'<AP)"""
    print("\n" + "="*80)
    print("4. 조치 후 리스크 검증 (Post-Action Risk Validation)")
    print("="*80)

    issues = []
    AP_order = {'L': 1, 'M': 2, 'H': 3}

    for i, row in enumerate(data, 1):
        S = row.get('S')
        O = row.get('O')
        D = row.get('D')
        AP = row.get('AP')
        S_prime = row.get("S'")
        O_prime = row.get("O'")
        D_prime = row.get("D'")
        AP_prime = row.get("AP'")

        prevention = row.get('예방조치', '')
        detection = row.get('검출조치', '')

        if not all(v is not None for v in [S, O, D, AP, S_prime, O_prime, D_prime, AP_prime]):
            continue

        try:
            S_int = int(S)
            O_int = int(O)
            D_int = int(D)
            S_prime_int = int(S_prime)
            O_prime_int = int(O_prime)
            D_prime_int = int(D_prime)

            # S' <= S 검증
            if S_prime_int > S_int:
                issues.append({
                    'row': i,
                    'type': "S' > S 오류",
                    'severity': 'ERROR',
                    'detail': f"S'={S_prime_int} > S={S_int} (심각도가 증가할 수 없음)"
                })

            # O' < O 검증 (예방조치가 있으면)
            if prevention and O_prime_int >= O_int:
                issues.append({
                    'row': i,
                    'type': "O' >= O 오류",
                    'severity': 'ERROR',
                    'detail': f"O'={O_prime_int} >= O={O_int} (예방조치 효과 없음: '{prevention[:30]}...')"
                })

            # D' < D 검증 (검출조치가 있으면)
            if detection and D_prime_int >= D_int:
                issues.append({
                    'row': i,
                    'type': "D' >= D 오류",
                    'severity': 'ERROR',
                    'detail': f"D'={D_prime_int} >= D={D_int} (검출조치 효과 없음: '{detection[:30]}...')"
                })

            # AP' < AP 검증
            if AP in AP_order and AP_prime in AP_order:
                if AP_order[AP_prime] >= AP_order[AP]:
                    issues.append({
                        'row': i,
                        'type': "AP' >= AP 경고",
                        'severity': 'WARNING',
                        'detail': f"AP'={AP_prime} >= AP={AP} (우선순위 개선 없음)"
                    })

        except (ValueError, TypeError) as e:
            issues.append({
                'row': i,
                'type': '값 형식 오류',
                'severity': 'ERROR',
                'detail': f"숫자 변환 실패: {e}"
            })

    if issues:
        print(f"\n[오류] 조치 후 리스크 이슈 발견: {len(issues)}건")
        errors = [x for x in issues if x['severity'] == 'ERROR']
        warnings = [x for x in issues if x['severity'] == 'WARNING']

        if errors:
            print(f"\n  [ERROR] 오류 ({len(errors)}건):")
            for issue in errors[:10]:
                print(f"    Row {issue['row']}: [{issue['type']}] {issue['detail']}")
            if len(errors) > 10:
                print(f"    ... 외 {len(errors) - 10}건")

        if warnings:
            print(f"\n  [WARN] 경고 ({len(warnings)}건):")
            for issue in warnings[:10]:
                print(f"    Row {issue['row']}: [{issue['type']}] {issue['detail']}")
            if len(warnings) > 10:
                print(f"    ... 외 {len(warnings) - 10}건")
    else:
        print("[OK] 조치 후 리스크 검증 통과")

    return issues

def generate_summary(data, all_issues):
    """검증 요약 보고서 생성"""
    print("\n" + "="*80)
    print("검증 요약 (Validation Summary)")
    print("="*80)

    total_rows = len(data)
    total_issues = sum(len(issues) for issues in all_issues.values())

    print(f"\n총 데이터 행: {total_rows}개")
    print(f"총 이슈 건수: {total_issues}건")

    print("\n분류별 이슈:")
    for category, issues in all_issues.items():
        errors = len([x for x in issues if x['severity'] == 'ERROR'])
        warnings = len([x for x in issues if x['severity'] == 'WARNING'])
        print(f"  {category}: {len(issues)}건 (오류: {errors}, 경고: {warnings})")

    # 품질 점수 계산
    error_count = sum(len([x for x in issues if x['severity'] == 'ERROR']) for issues in all_issues.values())
    warning_count = sum(len([x for x in issues if x['severity'] == 'WARNING']) for issues in all_issues.values())

    quality_score = max(0, 100 - (error_count * 5) - (warning_count * 2))

    print(f"\n품질 점수: {quality_score}/100")
    if quality_score >= 90:
        grade = "우수 (A)"
    elif quality_score >= 80:
        grade = "양호 (B)"
    elif quality_score >= 70:
        grade = "보통 (C)"
    elif quality_score >= 60:
        grade = "미흡 (D)"
    else:
        grade = "불량 (F)"

    print(f"품질 등급: {grade}")

    print("\n" + "="*80)

def main():
    filepath = 'c:/Users/jmyoo/.claude/skills/fmea-analysis/철심_FMEA.xlsx'

    print("="*80)
    print("철심_FMEA.xlsx 전체 검증")
    print("="*80)
    print(f"\n파일: {filepath}")

    # 1. 데이터 로드
    print("\n데이터 로드 중...")
    data, headers = load_fmea_data(filepath)
    data = expand_merged_cells(data)
    print(f"[v] {len(data)}개 행 로드 완료")
    print(f"[v] 컬럼: {', '.join(headers)}")

    # 2. 검증 실행
    all_issues = {}

    all_issues['인과관계'] = validate_causal_relationships(data)
    all_issues['리스크분석'] = validate_risk_ratings(data)
    all_issues['개선조치'] = validate_improvement_effectiveness(data)
    all_issues['조치후리스크'] = validate_post_action_risk(data)

    # 3. 요약 보고서
    generate_summary(data, all_issues)

if __name__ == "__main__":
    main()
